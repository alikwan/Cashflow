"""
app/api/export.py
==================
E1: On-demand report export endpoints.

  GET /api/export/excel  — multi-sheet .xlsx workbook from analytical tables
  GET /api/export/pdf    — one-page summary PDF

Both endpoints require a valid session cookie (get_current_user dependency).

PDF library chosen: fpdf2
  Rationale: WeasyPrint was attempted first (best Arabic/RTL via Pango/Cairo),
  but it fails to import on macOS because the native Pango/GObject system
  libraries are absent (OSError: cannot load library 'libgobject-2.0-0').
  fpdf2 renders a valid PDF with the Arabic font 'IBM Plex Sans Arabic'
  (IBMPlexSansArabic-Regular.ttf found in ~/Library/Fonts/) combined with
  arabic-reshaper + python-bidi for correct right-to-left ligature shaping.
  Full Arabic text is rendered correctly; RTL paragraph alignment is achieved
  via right-aligned cells.

NOTE: The Arabic font path is resolved at import time with a fallback to
Helvetica if the TTF is missing (production server may not have ~/Library/Fonts).
A future task should bundle the font inside the package for portability.
"""
from __future__ import annotations

import io
import os
from collections import defaultdict
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, get_session
from app.db.models import (
    Alert,
    BalancesSnapshot,
    ForecastBase,
    InstallmentsAging,
    InstallmentsSummary,
    MonthlyCashflow,
    Supplier,
    SupplierCap,
)

router = APIRouter(prefix="/api/export", tags=["export"])

# ---------------------------------------------------------------------------
# Arabic text helper (reshaper + bidi → correct PDF glyphs)
# ---------------------------------------------------------------------------

try:
    import arabic_reshaper
    from bidi.algorithm import get_display as _bidi_display

    def _ar(text: str) -> str:
        """Reshape Arabic text for correct RTL ligature rendering in PDF."""
        return _bidi_display(arabic_reshaper.reshape(text))

    _ARABIC_OK = True
except ImportError:
    # Fallback: return text as-is (glyphs will be wrong but PDF is still valid)
    def _ar(text: str) -> str:  # type: ignore[misc]
        return text

    _ARABIC_OK = False


# ---------------------------------------------------------------------------
# Arabic font path resolution
# ---------------------------------------------------------------------------

_FONT_CANDIDATES = [
    os.path.expanduser("~/Library/Fonts/IBMPlexSansArabic-Regular.ttf"),
    "/usr/share/fonts/truetype/IBMPlexSansArabic-Regular.ttf",
    os.path.expanduser("~/Library/Fonts/NotoNaskhArabic-Regular.ttf"),
    "/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf",
]

_ARABIC_FONT_PATH: Optional[str] = None
for _p in _FONT_CANDIDATES:
    if os.path.isfile(_p):
        _ARABIC_FONT_PATH = _p
        break


# ---------------------------------------------------------------------------
# Excel workbook builder
# ---------------------------------------------------------------------------

def build_workbook(db: Session) -> io.BytesIO:
    """
    Build an xlsx workbook from the analytical tables.

    Sheets (Arabic titles, Unicode handled natively by openpyxl):
      1. التدفق الشهري   — monthly_cashflow rows
      2. الموردون         — suppliers + latest cap + latest balance
      3. الأقساط          — installments_summary + aging buckets
      4. التنبؤ           — forecast_base pivoted (series × month)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill("solid", fgColor="2563EB")  # brand blue
    HEADER_ALIGN = Alignment(horizontal="center")
    NUM_FMT = "#,##0.0"

    def _style_header(ws, headers: list[str]) -> None:
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        ws.freeze_panes = "A2"

    def _set_col_widths(ws, widths: list[int]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _apply_num_fmt(ws, col: int, start_row: int = 2, fmt: str = NUM_FMT) -> None:
        """Apply number format to an entire column (from start_row down)."""
        for row in ws.iter_rows(min_row=start_row, min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = fmt

    # ------------------------------------------------------------------
    # Sheet 1: التدفق الشهري (Monthly Cashflow)
    # ------------------------------------------------------------------
    ws1 = wb.create_sheet("التدفق الشهري")
    headers1 = [
        "الشهر", "السنة المالية",
        "مقبوضات (م)", "مدفوعات موردين (م)", "سحوبات (م)",
        "مرتجعات (م)", "مشتريات (م)", "أجور (م)", "صيرفة (م)", "أخرى (م)",
        "إجمالي خروج (م)", "صافي (م)", "الرصيد التراكمي (م)",
    ]
    _style_header(ws1, headers1)
    _set_col_widths(ws1, [10, 12, 14, 18, 12, 12, 12, 10, 10, 10, 16, 12, 18])

    cf_rows = (
        db.query(MonthlyCashflow)
        .order_by(MonthlyCashflow.year_month.asc())
        .all()
    )
    for r, row in enumerate(cf_rows, start=2):
        ws1.cell(r, 1, row.year_month)
        ws1.cell(r, 2, row.fiscal_year)
        ws1.cell(r, 3, float(row.cash_in_m))
        ws1.cell(r, 4, float(row.out_suppliers_m))
        ws1.cell(r, 5, float(row.out_drawings_m))
        ws1.cell(r, 6, float(row.out_refunds_m))
        ws1.cell(r, 7, float(row.out_purchases_m))
        ws1.cell(r, 8, float(row.out_salaries_m))
        ws1.cell(r, 9, float(row.out_siyrafa_m))
        ws1.cell(r, 10, float(row.out_other_m))
        ws1.cell(r, 11, float(row.out_total_comprehensive_m))
        ws1.cell(r, 12, float(row.net_total_m))
        ws1.cell(r, 13, float(row.cash_running_m))

    for col_idx in range(3, 14):
        _apply_num_fmt(ws1, col_idx)

    # ------------------------------------------------------------------
    # Sheet 2: الموردون (Suppliers)
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("الموردون")
    headers2 = [
        "معرّف الحساب", "الاسم", "العملة", "سقف شهري (م)",
        "الرصيد الحالي (م)", "الرصيد بالدينار (م)", "نشط",
    ]
    _style_header(ws2, headers2)
    _set_col_widths(ws2, [14, 22, 8, 16, 18, 20, 8])

    suppliers = (
        db.query(Supplier)
        .order_by(Supplier.display_order.asc(), Supplier.account_id.asc())
        .all()
    )

    # Latest cap per supplier_id
    caps: dict[int, float] = {}
    for sup in suppliers:
        cap_row = (
            db.query(SupplierCap)
            .filter(SupplierCap.supplier_id == sup.id)
            .order_by(SupplierCap.effective_from.desc())
            .first()
        )
        caps[sup.id] = float(cap_row.monthly_cap_m) if cap_row else 0.0

    # Latest balance snapshot per account_id (sum over currency_ids → IQD)
    bal_snap_date: Optional[date] = None
    snap_q = db.query(BalancesSnapshot).order_by(BalancesSnapshot.snapshot_date.desc()).first()
    if snap_q:
        bal_snap_date = snap_q.snapshot_date

    balances_iqd: dict[int, float] = defaultdict(float)
    balances_raw: dict[int, float] = defaultdict(float)
    if bal_snap_date:
        snap_rows = (
            db.query(BalancesSnapshot)
            .filter(BalancesSnapshot.snapshot_date == bal_snap_date)
            .all()
        )
        for sn in snap_rows:
            balances_iqd[sn.account_id] += float(sn.balance_iqd_m)
            balances_raw[sn.account_id] += float(sn.balance_m)

    for r, sup in enumerate(suppliers, start=2):
        ws2.cell(r, 1, sup.account_id)
        ws2.cell(r, 2, sup.name)
        ws2.cell(r, 3, sup.currency or "IQD")
        ws2.cell(r, 4, caps.get(sup.id, 0.0))
        ws2.cell(r, 5, balances_raw.get(sup.account_id, 0.0))
        ws2.cell(r, 6, balances_iqd.get(sup.account_id, 0.0))
        ws2.cell(r, 7, "نعم" if sup.active else "لا")

    for col_idx in (4, 5, 6):
        _apply_num_fmt(ws2, col_idx)

    # ------------------------------------------------------------------
    # Sheet 3: الأقساط (Installments)
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("الأقساط")

    # Summary section
    inst = (
        db.query(InstallmentsSummary)
        .order_by(InstallmentsSummary.snapshot_date.desc())
        .first()
    )

    summary_headers = [
        "البند", "القيمة (مليون دينار)",
    ]
    _style_header(ws3, summary_headers)
    _set_col_widths(ws3, [30, 22])

    summary_rows = []
    if inst:
        summary_rows = [
            ("عدد الأقساط", int(inst.premium_count)),
            ("القيمة الاسمية الإجمالية", float(inst.face_total_m)),
            ("النقد المُحصَّل", float(inst.cash_paid_m)),
            ("الخصومات", float(inst.discount_m)),
            ("الرصيد القائم", float(inst.remaining_m)),
        ]
    else:
        summary_rows = [
            ("عدد الأقساط", 0),
            ("القيمة الاسمية الإجمالية", 0.0),
            ("النقد المُحصَّل", 0.0),
            ("الخصومات", 0.0),
            ("الرصيد القائم", 0.0),
        ]

    for r, (label, val) in enumerate(summary_rows, start=2):
        ws3.cell(r, 1, label)
        cell = ws3.cell(r, 2, val)
        if isinstance(val, float):
            cell.number_format = NUM_FMT

    # Aging section (skip a row, then subheader)
    aging_start = len(summary_rows) + 3  # +2 for data rows, +1 gap

    ws3.cell(aging_start, 1, "أعمار الديون").font = Font(bold=True)
    aging_start += 1
    ws3.cell(aging_start, 1, "الفئة")
    ws3.cell(aging_start, 2, "المبلغ (مليون دينار)")
    ws3.cell(aging_start, 3, "العدد")
    for col in (1, 2, 3):
        cell = ws3.cell(aging_start, col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    _set_col_widths(ws3, [30, 22, 12])

    snap_date_inst: Optional[date] = inst.snapshot_date if inst else None
    aging_rows = []
    if snap_date_inst:
        aging_rows = (
            db.query(InstallmentsAging)
            .filter(InstallmentsAging.snapshot_date == snap_date_inst)
            .order_by(InstallmentsAging.bucket_key.asc())
            .all()
        )

    for r, ag in enumerate(aging_rows, start=aging_start + 1):
        ws3.cell(r, 1, ag.label or ag.bucket_key)
        cell_amt = ws3.cell(r, 2, float(ag.amount_m))
        cell_amt.number_format = NUM_FMT
        ws3.cell(r, 3, ag.count)

    # ------------------------------------------------------------------
    # Sheet 4: التنبؤ (Forecast)
    # ------------------------------------------------------------------
    ws4 = wb.create_sheet("التنبؤ")

    # Pivot: rows = series_key, cols = year_month (for engine='seasonal')
    fc_rows = (
        db.query(ForecastBase)
        .filter(ForecastBase.engine == "seasonal")
        .order_by(ForecastBase.series_key.asc(), ForecastBase.year_month.asc())
        .all()
    )

    # Collect unique months (ordered) and series
    months_seen: list[str] = []
    months_set: set[str] = set()
    series_data: dict[str, dict[str, float]] = defaultdict(dict)
    for fc in fc_rows:
        if fc.year_month not in months_set:
            months_seen.append(fc.year_month)
            months_set.add(fc.year_month)
        series_data[fc.series_key][fc.year_month] = float(fc.value_m)

    fc_headers = ["السلسلة"] + months_seen
    _style_header(ws4, fc_headers)
    widths4 = [18] + [10] * len(months_seen)
    _set_col_widths(ws4, widths4)

    series_labels: dict[str, str] = {
        "cash_in": "مقبوضات",
        "out_suppliers": "موردون",
        "out_drawings": "سحوبات",
        "out_refunds": "مرتجعات",
        "out_purchases": "مشتريات",
        "out_salaries": "أجور",
        "out_siyrafa": "صيرفة",
        "out_other": "أخرى",
    }

    for r, (skey, mdict) in enumerate(sorted(series_data.items()), start=2):
        ws4.cell(r, 1, series_labels.get(skey, skey))
        for c, ym in enumerate(months_seen, start=2):
            cell = ws4.cell(r, c, mdict.get(ym, 0.0))
            cell.number_format = NUM_FMT

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# PDF summary builder
# ---------------------------------------------------------------------------

def build_summary_pdf(db: Session) -> bytes:
    """
    Build a one-page summary PDF using fpdf2 with IBM Plex Sans Arabic font.

    Arabic text is correctly reshaped via arabic-reshaper + python-bidi.
    Layout: RTL, right-aligned, numeric columns use Latin digits (standard
    for financial data in Iraq). If the Arabic TTF is not found on this
    machine, Helvetica is used and Arabic glyphs will be missing (Latin labels
    used as fallback).
    """
    from fpdf import FPDF, XPos, YPos

    # ------------------------------------------------------------------
    # Gather data
    # ------------------------------------------------------------------
    last_cf = (
        db.query(MonthlyCashflow)
        .order_by(MonthlyCashflow.year_month.desc())
        .first()
    )
    current_cash_m: float = float(last_cf.cash_running_m) if last_cf else 0.0

    inst = (
        db.query(InstallmentsSummary)
        .order_by(InstallmentsSummary.snapshot_date.desc())
        .first()
    )
    face_m = float(inst.face_total_m) if inst else 0.0
    remaining_m = float(inst.remaining_m) if inst else 0.0
    cash_paid_m = float(inst.cash_paid_m) if inst else 0.0

    # FY totals: aggregate all monthly rows
    cf_rows = db.query(MonthlyCashflow).all()
    fy_sums: dict[str, dict[str, float]] = defaultdict(lambda: {"in_m": 0.0, "out_m": 0.0})
    for r in cf_rows:
        fy_sums[r.fiscal_year]["in_m"] += float(r.cash_in_m)
        fy_sums[r.fiscal_year]["out_m"] += float(r.out_total_comprehensive_m)

    # Active alerts count
    active_alerts = (
        db.query(Alert)
        .filter(Alert.status != "resolved")
        .count()
    )

    # Top 3 suppliers by display_order
    top_suppliers = (
        db.query(Supplier)
        .filter(Supplier.active == True)  # noqa: E712
        .order_by(Supplier.display_order.asc(), Supplier.account_id.asc())
        .limit(3)
        .all()
    )

    # Report date
    today = date.today().isoformat()

    # ------------------------------------------------------------------
    # Build PDF
    # ------------------------------------------------------------------
    pdf = FPDF()
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()

    use_arabic = _ARABIC_FONT_PATH is not None and _ARABIC_OK

    if use_arabic:
        pdf.add_font("Arabic", "", _ARABIC_FONT_PATH)
        pdf.add_font("Arabic", "B", _ARABIC_FONT_PATH)

    def set_arabic_font(size: int, bold: bool = False) -> None:
        if use_arabic:
            style = "B" if bold else ""
            pdf.set_font("Arabic", style, size)
        else:
            pdf.set_font("Helvetica", "B" if bold else "", size)

    def rtext(text: str) -> str:
        """Return text ready for PDF rendering (reshaped + bidi if Arabic)."""
        if use_arabic and _ARABIC_OK:
            return _ar(text)
        return text

    def row(label_ar: str, value: str, label_fallback: str = "") -> None:
        """Render a label+value line, right-aligned."""
        label = rtext(label_ar) if use_arabic else label_fallback or label_ar
        set_arabic_font(10)
        pdf.set_x(pdf.l_margin)
        w = pdf.w - pdf.l_margin - pdf.r_margin
        pdf.cell(w * 0.55, 8, value, align="L")
        pdf.cell(w * 0.45, 8, label, align="R",
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    def divider() -> None:
        pdf.set_draw_color(37, 99, 235)  # brand blue
        x = pdf.l_margin
        y = pdf.get_y() + 2
        pdf.line(x, y, pdf.w - pdf.r_margin, y)
        pdf.set_y(y + 4)

    # ------------------------------------------------------------------
    # Title
    # ------------------------------------------------------------------
    set_arabic_font(18, bold=True)
    pdf.set_text_color(37, 99, 235)
    pdf.cell(0, 14, rtext("تقرير السيولة النقدية — معرض البيت السعيد"),
             align="R", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(0, 0, 0)

    set_arabic_font(9)
    pdf.cell(0, 6, today, align="R", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(2)

    # ------------------------------------------------------------------
    # Section 1: Current cash position
    # ------------------------------------------------------------------
    set_arabic_font(12, bold=True)
    pdf.cell(0, 10, rtext("الوضع النقدي الحالي"), align="R",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    divider()

    set_arabic_font(10)
    row(
        "الرصيد النقدي التراكمي (مليون د.ع)",
        f"{current_cash_m:,.1f}",
        label_fallback="Current cash (M IQD)",
    )
    if last_cf:
        row(
            "آخر شهر مُحلَّل",
            last_cf.year_month,
            label_fallback="Last analysed month",
        )
    pdf.ln(3)

    # ------------------------------------------------------------------
    # Section 2: FY totals
    # ------------------------------------------------------------------
    set_arabic_font(12, bold=True)
    pdf.cell(0, 10, rtext("ملخّص السنوات المالية"), align="R",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    divider()

    set_arabic_font(9)
    for fy, vals in sorted(fy_sums.items()):
        net = vals["in_m"] - vals["out_m"]
        row(
            f"{rtext('السنة المالية')} {fy}",
            f"IN {vals['in_m']:,.1f}  |  OUT {vals['out_m']:,.1f}  |  NET {net:,.1f}",
            label_fallback=f"FY {fy}",
        )
    if not fy_sums:
        row("السنوات المالية", "لا توجد بيانات", label_fallback="FY totals: no data")
    pdf.ln(3)

    # ------------------------------------------------------------------
    # Section 3: Installments summary
    # ------------------------------------------------------------------
    set_arabic_font(12, bold=True)
    pdf.cell(0, 10, rtext("ملخّص الأقساط"), align="R",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    divider()

    set_arabic_font(10)
    row("القيمة الاسمية (مليون)", f"{face_m:,.1f}", "Face total (M)")
    row("النقد المُحصَّل (مليون)", f"{cash_paid_m:,.1f}", "Cash paid (M)")
    row("الرصيد القائم (مليون)", f"{remaining_m:,.1f}", "Remaining (M)")
    collection_rate = (cash_paid_m / face_m * 100) if face_m > 0 else 0.0
    row("نسبة التحصيل", f"{collection_rate:.1f}%", "Collection rate")
    pdf.ln(3)

    # ------------------------------------------------------------------
    # Section 4: Alerts + top suppliers
    # ------------------------------------------------------------------
    set_arabic_font(12, bold=True)
    pdf.cell(0, 10, rtext("التنبيهات والموردون"), align="R",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    divider()

    set_arabic_font(10)
    row("التنبيهات النشطة", str(active_alerts), "Active alerts")
    pdf.ln(2)

    set_arabic_font(10, bold=True)
    pdf.cell(0, 7, rtext("أبرز الموردين:"), align="R",
             new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    set_arabic_font(10)
    for sup in top_suppliers:
        name_display = rtext(sup.name) if use_arabic else sup.name
        row(name_display, sup.currency or "IQD", label_fallback=sup.name)

    # ------------------------------------------------------------------
    # Footer
    # ------------------------------------------------------------------
    pdf.set_y(-20)
    set_arabic_font(8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 6,
             rtext("تقرير مُولَّد تلقائياً — نظام إدارة السيولة النقدية"),
             align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("/excel")
def export_excel(
    db: Session = Depends(get_session),
    _user=Depends(get_current_user),
) -> StreamingResponse:
    """Export the full analytical dataset as a multi-sheet .xlsx workbook."""
    buf = build_workbook(db)
    filename = f"cashflow_report_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get("/pdf")
def export_pdf(
    db: Session = Depends(get_session),
    _user=Depends(get_current_user),
) -> Response:
    """Export a one-page summary PDF."""
    pdf_bytes = build_summary_pdf(db)
    filename = f"cashflow_summary_{date.today().isoformat()}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )
