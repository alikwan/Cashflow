"""
tests/api/test_export.py
========================
TDD tests for E1: GET /api/export/excel and GET /api/export/pdf.

Both endpoints:
  - require auth (401 without session cookie)
  - return correct Content-Type and Content-Disposition
  - work on EMPTY DB (no seed_analytics) without raising 500
  - return valid file magic bytes
  - (xlsx) contain expected sheet names
"""
from __future__ import annotations

import io
import pytest
import openpyxl


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

_EXPECTED_SHEETS = {
    "التدفق الشهري",
    "الموردون",
    "الأقساط",
    "التنبؤ",
}


# ---------------------------------------------------------------------------
# Auth guard — no cookie → 401
# ---------------------------------------------------------------------------

def test_export_excel_requires_auth(client):
    r = client.get("/api/export/excel")
    assert r.status_code == 401


def test_export_pdf_requires_auth(client):
    r = client.get("/api/export/pdf")
    assert r.status_code == 401


# ---------------------------------------------------------------------------
# Excel — happy path (with seed data)
# ---------------------------------------------------------------------------

def test_export_excel_returns_xlsx(client, seed_analytics, auth):
    r = client.get("/api/export/excel", cookies=auth)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # xlsx is a ZIP → starts with PK magic
    assert r.content[:2] == b"PK"
    assert "attachment" in r.headers.get("content-disposition", "")
    assert "cashflow_report_" in r.headers.get("content-disposition", "")


def test_export_excel_sheet_names(client, seed_analytics, auth):
    r = client.get("/api/export/excel", cookies=auth)
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    sheet_names = set(wb.sheetnames)
    assert _EXPECTED_SHEETS.issubset(sheet_names), (
        f"Missing sheets: {_EXPECTED_SHEETS - sheet_names}"
    )


def test_export_excel_monthly_sheet_has_data(client, seed_analytics, auth):
    """The monthly cashflow sheet should have at least a header + 24 data rows."""
    r = client.get("/api/export/excel", cookies=auth)
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb["التدفق الشهري"]
    # header row + 24 seed rows
    assert ws.max_row >= 25


# ---------------------------------------------------------------------------
# Excel — empty DB (no seed_analytics) must not 500
# ---------------------------------------------------------------------------

def test_export_excel_empty_db(client, auth):
    """Excel export on an empty analytical DB must succeed (empty sheets, not 500)."""
    r = client.get("/api/export/excel", cookies=auth)
    assert r.status_code == 200
    # still a valid xlsx
    assert r.content[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert _EXPECTED_SHEETS.issubset(set(wb.sheetnames))


# ---------------------------------------------------------------------------
# PDF — happy path (with seed data)
# ---------------------------------------------------------------------------

def test_export_pdf_returns_pdf(client, seed_analytics, auth):
    r = client.get("/api/export/pdf", cookies=auth)
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"
    assert r.headers["content-type"].startswith("application/pdf")
    assert "attachment" in r.headers.get("content-disposition", "")
    assert "cashflow_summary_" in r.headers.get("content-disposition", "")


# ---------------------------------------------------------------------------
# PDF — empty DB must not 500
# ---------------------------------------------------------------------------

def test_export_pdf_empty_db(client, auth):
    """PDF export on an empty analytical DB must succeed (summary with zeros)."""
    r = client.get("/api/export/pdf", cookies=auth)
    assert r.status_code == 200
    assert r.content[:4] == b"%PDF"


# ---------------------------------------------------------------------------
# PDF — no-font fallback must not crash (I3)
# ---------------------------------------------------------------------------

def test_export_pdf_without_arabic_font(client, seed_analytics, auth, monkeypatch):
    """
    When BOTH the Arabic font is missing (_ARABIC_FONT_PATH=None) AND the
    reshaper/bidi libraries are unavailable (_ARABIC_OK=False), the PDF builder
    must fall back to Latin-only Helvetica text and return a valid PDF — it must
    NOT raise FPDFUnicodeEncodingException (which would surface as a 500).

    This is the full worst-case fallback: the real fix is bundling the font
    (app/api/fonts/IBMPlexSansArabic-Regular.ttf), but the code must stay
    crash-safe even if neither the font nor the shaping libraries are present.
    """
    import app.api.export as export_module

    # No bundled/system font found AND no arabic-reshaper/bidi available.
    monkeypatch.setattr(export_module, "_ARABIC_FONT_PATH", None)
    monkeypatch.setattr(export_module, "_ARABIC_OK", False)

    r = client.get("/api/export/pdf", cookies=auth)
    assert r.status_code == 200, (
        f"PDF export crashed in Helvetica fallback (status={r.status_code}): "
        f"{r.text[:500]}"
    )
    assert r.content[:4] == b"%PDF", "Response is not a valid PDF"


def test_export_pdf_without_arabic_font_empty_db(client, auth, monkeypatch):
    """Helvetica fallback must also be crash-safe on an EMPTY analytical DB.

    Empty DB exercises the zero-value / 'no data' label branches (e.g. the
    'FY totals: no data' row), which must still avoid sending raw Arabic to
    Helvetica.
    """
    import app.api.export as export_module

    monkeypatch.setattr(export_module, "_ARABIC_FONT_PATH", None)
    monkeypatch.setattr(export_module, "_ARABIC_OK", False)

    r = client.get("/api/export/pdf", cookies=auth)
    assert r.status_code == 200, (
        f"PDF export crashed in Helvetica fallback on empty DB "
        f"(status={r.status_code}): {r.text[:500]}"
    )
    assert r.content[:4] == b"%PDF", "Response is not a valid PDF"
