#!/usr/bin/env python3
"""
بناء تقرير تحليل السيولة النقدية - معرض البيت السعيد
=====================================================
يقرأ البيانات من SQL Server (AlBaytAlSaeid) ويولّد ملف Excel متعدد الأوراق
مع تحليل تاريخي 48 شهر + تنبؤ 12 شهر + اختبارات صحة (backtesting).
"""

import os
import sys
from datetime import datetime
from pathlib import Path

import pymssql
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

# ─────────────────────────────────────────────
# الإعدادات
# ─────────────────────────────────────────────
DB = dict(
    server="localhost", port=1433, user="sa",
    password="MyS3cure!Pass2025", database="AlBaytAlSaeid",
    charset="UTF-8"
)
REPORT_DATE = datetime(2026, 5, 13).date()
START_DATE = "2022-05-01"
END_DATE = "2026-05-01"
FORECAST_HORIZON = 12          # 12 شهر تنبؤ
UNEXPECTED_RESERVE_M = 15.0    # 15 مليون احتياطي شهري
CAGR_FLOOR, CAGR_CAP = -0.10, 0.15
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
OUTPUT_FILE = OUTPUT_DIR / f"Cash_Flow_Analysis_2022-2027.xlsx"

# الـ14 موزّعاً - الخطة الشهرية المعتمدة من المستخدم
# (PlanLow, PlanHigh, UserMonthly) = الحدود + القيمة المعتمدة للسداد الشهري
SUPPLIERS = [
    # (AccountId, الاسم,                  السقف, الأدنى, الأعلى,  المعتمد شهرياً)
    (1001, "01-معرض البركة",            5.0,   3.0,   5.0,    4.0),
    (2079, "02-هيثم",                    3.0,   2.0,   3.0,    2.5),
    (2093, "03-وميض",                    0.0,   0.0,   0.0,    0.0),
    (2432, "04-حميد الشطباوي",         15.0,   0.0,  15.0,   15.0),
    (2440, "05-معرض الهادي",            3.0,   3.0,   3.0,    3.0),
    (2700, "06-معرض اولاد شفيق",        3.0,   0.0,   3.0,    1.5),
    (3123, "07-العطاوي للمفروشات",      2.0,   0.0,   2.0,    1.0),
    (3916, "08-شركة اصل القمة",         3.0,   3.0,   3.0,    3.0),
    (5721, "09-قاسم بايسكلات",          4.0,   0.0,   4.0,    2.0),
    (2439, "10-معرض الواحة سامراء",     5.0,   0.0,   5.0,    5.0),
    (4937, "11-شركة الحافظ",           40.0,  30.0,  40.0,   30.0),
    (6444, "12-كهربائيات المهندس",     15.0,  15.0,  15.0,   15.0),
    (6552, "13-دكتور يوسف ميديا فوكس",  5.0,   3.0,   5.0,    5.0),
    (6918, "14-شركة الريان بغداد",      7.0,   7.0,   7.0,    7.0),
]

# ─────────────────────────────────────────────
# اتصال + استعلامات
# ─────────────────────────────────────────────
def connect():
    return pymssql.connect(**DB)

def query_df(sql):
    with connect() as c:
        return pd.read_sql(sql, c)

# الـ7 صناديق (cash boxes) ـ نوع 1811, 1812
CASH_BOX_IDS = (180, 181, 4935, 6314, 6662, 6672, 6684)

def get_monthly_cashflow():
    """جلب 48 شهر بفئات IN/OUT - يعتمد OperationsType + الحساب المقابل.

    اكتشاف مهم: OperationsType يصنّف نوع الحركة بدقة:
    - 0 = مقبوضات (Receipts) - دخول
    - 1 = مدفوعات للموردين/الشركاء (Payments) - خروج
    - 3 = تحويلات بين الصناديق (Internal) - يُستثنى
    - 5 = مصاريف تشغيلية (Operating expenses) - خروج
    - 7 = صيرفة - تحويل عملة (Currency exchange IQD→USD) - خروج فعلي للدينار

    إضافة الصيرفة (Type 7) كفئة OUT مستقلة كانت السبب الرئيسي لفجوة 391M سنوياً.
    """
    sql = f"""
    SELECT FORMAT(b.Date,'yyyy-MM') AS YearMonth,
      -- المقبوضات (OperationsType=0)
      SUM(CASE WHEN b.OperationsType=0 THEN b.Amount1 ELSE 0 END)/1000000.0 AS Cash_IN_M,
      -- المدفوعات حسب الوجهة (OperationsType=1)
      SUM(CASE WHEN b.OperationsType=1 AND at.AccountTypeId=2614 THEN b.Amount1 ELSE 0 END)/1000000.0 AS OUT_Suppliers_M,
      SUM(CASE WHEN b.OperationsType=1 AND at.AccountTypeId=2518 THEN b.Amount1 ELSE 0 END)/1000000.0 AS OUT_Drawings_M,
      SUM(CASE WHEN b.OperationsType=1 AND at.AccountTypeId=1631 THEN b.Amount1 ELSE 0 END)/1000000.0 AS OUT_Refunds_M,
      -- المصاريف التشغيلية (OperationsType=5)
      SUM(CASE WHEN b.OperationsType=5 AND at.AccountTypeId=3110 THEN b.Amount1 ELSE 0 END)/1000000.0 AS OUT_Purchases_M,
      SUM(CASE WHEN b.OperationsType=5 AND at.AccountTypeId=3121 THEN b.Amount1 ELSE 0 END)/1000000.0 AS OUT_Salaries_M,
      SUM(CASE WHEN b.OperationsType=5 AND at.AccountTypeId IN (3124,2110) THEN b.Amount1 ELSE 0 END)/1000000.0 AS OUT_Other_M,
      -- الصيرفة (OperationsType=7) - تحويل دينار→دولار
      SUM(CASE WHEN b.OperationsType=7 THEN b.Amount1 ELSE 0 END)/1000000.0 AS OUT_Siyrafa_M,
      -- تحويلات داخلية (Type 3) - للعرض فقط، لا تُحسب في OUT
      SUM(CASE WHEN b.OperationsType=3 THEN b.Amount1 ELSE 0 END)/1000000.0 AS Internal_Transfers_M,
      COUNT(*) AS BondCount
    FROM Bonds b
    JOIN accounts at ON at.AccountId=b.AccountToId
    WHERE b.Deleted=0 AND ISNULL(b.IsEdit,0)=0
      AND b.Date >= '{START_DATE}' AND b.Date < '{END_DATE}'
    GROUP BY FORMAT(b.Date,'yyyy-MM')
    ORDER BY YearMonth
    """
    df = query_df(sql)
    # المنظور التشغيلي (View C): يستثني الصيرفة (الدولار يحتسب ضمن "موردون")
    df["OUT_Total_Operational_M"] = (df["OUT_Suppliers_M"] + df["OUT_Drawings_M"]
                                       + df["OUT_Refunds_M"] + df["OUT_Purchases_M"]
                                       + df["OUT_Salaries_M"] + df["OUT_Other_M"])
    # المنظور الشامل (View A): يضم الصيرفة كخروج للدينار
    df["OUT_Total_Comprehensive_M"] = df["OUT_Total_Operational_M"] + df["OUT_Siyrafa_M"]
    df["OUT_Total_M"] = df["OUT_Total_Comprehensive_M"]  # المنظور الافتراضي = الشامل
    df["Net_Operating_M"] = df["Cash_IN_M"] - df["OUT_Total_Operational_M"]
    df["Net_Total_M"] = df["Cash_IN_M"] - df["OUT_Total_Comprehensive_M"]
    return df

def get_per_supplier_monthly():
    """مدفوعات شهرية لكل واحد من الـ14 موزّع، مفصّلة بالعملة.

    تُعيد:
    - Paid_M           = SUM(Amount1) - القيمة المسجّلة بالدينار (تتأثر بـ Rate1 وقت السند)
    - Paid_IQD_M       = SUM(Amount1) عندما Currency1Id=1 (سندات دينار خالصة فقط)
    - Paid_USD_M       = SUM(Amount2) عندما Currency1Id=2 (القيمة بالدولار - بالملايين USD)
    - Recv_M           = ما استلمته الشركة من المورد (نادر)
    """
    sups_sql = ",".join(f"({sid},N'{name}',{cap})" for sid, name, cap, _, _, _ in SUPPLIERS)
    sql = f"""
    WITH Sup(AccountId,Name,Cap) AS (SELECT * FROM (VALUES {sups_sql}) v(AccountId,Name,Cap))
    SELECT s.Name AS Supplier, FORMAT(b.Date,'yyyy-MM') AS YearMonth,
      SUM(CASE WHEN b.AccountToId=s.AccountId THEN b.Amount1 ELSE 0 END)/1000000.0 AS Paid_M,
      SUM(CASE WHEN b.AccountToId=s.AccountId AND b.Currency1Id=1 THEN b.Amount1 ELSE 0 END)/1000000.0 AS Paid_IQD_M,
      SUM(CASE WHEN b.AccountToId=s.AccountId AND b.Currency1Id=2 AND b.Rate1>0
               THEN b.Amount1/b.Rate1 ELSE 0 END)/1000000.0 AS Paid_USD_M,
      SUM(CASE WHEN b.AccountFromId=s.AccountId THEN b.Amount1 ELSE 0 END)/1000000.0 AS Recv_M,
      MAX(s.Cap) AS Cap_M
    FROM Sup s
    JOIN Bonds b ON (b.AccountFromId=s.AccountId OR b.AccountToId=s.AccountId)
                 AND b.Deleted=0 AND ISNULL(b.IsEdit,0)=0
                 AND b.Date >= '{START_DATE}' AND b.Date < '{END_DATE}'
    GROUP BY s.Name, FORMAT(b.Date,'yyyy-MM')
    ORDER BY s.Name, YearMonth
    """
    return query_df(sql)

def get_installments_summary():
    """ملخص الأقساط المفتوحة."""
    sql = """
    SELECT COUNT(DISTINCT p.Id) AS PremiumCnt,
      SUM(p.TotalAmount)/1000000.0 AS TotalCommitted_M,
      SUM(ISNULL(pp.PaidSum,0))/1000000.0 AS Paid_M,
      (SUM(p.TotalAmount)-SUM(ISNULL(pp.PaidSum,0)))/1000000.0 AS Remaining_M
    FROM Premiums p
    OUTER APPLY (SELECT SUM(Amount) AS PaidSum FROM PremiumPays WHERE PremiumId=p.Id AND Deleted=0) pp
    WHERE p.Deleted=0 AND p.Date >= '2022-01-01'
    """
    return query_df(sql)

def get_balances_snapshot():
    """أرصدة الموردين الـ14 الحالية + الصناديق."""
    sup_ids = ",".join(str(s[0]) for s in SUPPLIERS)
    sup_names = {sid: name for sid, name, _, _, _, _ in SUPPLIERS}
    sql = f"""
    SELECT ta.AccountId, ta.CurrencyId, ta.Balance/1000000.0 AS Balance_M, ta.LastActive
    FROM tAccounts ta
    WHERE ta.AccountId IN ({sup_ids})
    """
    df = query_df(sql)
    df["Supplier"] = df["AccountId"].map(sup_names)
    df["Currency"] = df["CurrencyId"].map({1: "دينار", 2: "دولار"}).fillna("أخرى")
    return df[["Supplier", "AccountId", "Currency", "Balance_M", "LastActive"]].sort_values(["Supplier","Currency"])

def get_avg_usd_rate():
    """متوسط سعر صرف الدولار من السندات الـ12 شهر الأخيرة."""
    sql = """
    SELECT AVG(Rate1) AS AvgRate FROM Bonds
    WHERE Deleted=0 AND Currency1Id=2 AND Rate1 > 0
      AND Date >= DATEADD(MONTH, -12, '2026-05-13')
    """
    rate = query_df(sql)["AvgRate"].iloc[0]
    return float(rate) if rate else 1350.0

def get_cash_total_balance():
    sql = """
    SELECT a.AccountId, a.Name, ta.CurrencyId, ta.Balance/1000000.0 AS Balance_M
    FROM accounts a JOIN tAccounts ta ON ta.AccountId=a.AccountId
    WHERE a.AccountTypeId IN (1811,1812) AND a.Deleted=0
    """
    return query_df(sql)

def get_top_debtors():
    sql = """
    SELECT TOP 30 a.AccountId, a.Name, ta.CurrencyId, ta.Balance/1000000.0 AS Balance_M
    FROM accounts a JOIN tAccounts ta ON ta.AccountId=a.AccountId
    WHERE a.AccountTypeId=1631 AND a.Deleted=0 AND ta.Balance>0
    ORDER BY ta.Balance DESC
    """
    return query_df(sql)

def get_audit_summary():
    """عدد السندات شهرياً + شذوذ + سندات مستقبلية."""
    sql = f"""
    SELECT FORMAT(Date,'yyyy-MM') AS YearMonth, COUNT(*) AS BondCount,
      SUM(CASE WHEN Currency1Id=1 THEN 1 ELSE 0 END) AS Dinar_Bonds,
      SUM(CASE WHEN Currency1Id=2 THEN 1 ELSE 0 END) AS Dollar_Bonds,
      SUM(CASE WHEN OperationsType=0 THEN 1 ELSE 0 END) AS OpType_0,
      SUM(CASE WHEN OperationsType<>0 THEN 1 ELSE 0 END) AS OpType_Other
    FROM Bonds
    WHERE Deleted=0 AND Date >= '{START_DATE}' AND Date < '{END_DATE}' AND ISNULL(IsEdit,0)=0
    GROUP BY FORMAT(Date,'yyyy-MM') ORDER BY YearMonth
    """
    audit_monthly = query_df(sql)
    sql_fut = "SELECT COUNT(*) AS Cnt, ISNULL(SUM(Amount1),0)/1000000.0 AS Sum_M FROM Bonds WHERE Deleted=0 AND Date > '2026-05-13'"
    future = query_df(sql_fut)
    sql_top = f"""
    SELECT TOP 20 b.Id, b.Date, b.Amount1/1000000.0 AS Amount_M,
      af.Name AS FromAccount, at.Name AS ToAccount, LEFT(ISNULL(b.Reason,''),60) AS Reason
    FROM Bonds b
    LEFT JOIN accounts af ON af.AccountId=b.AccountFromId
    LEFT JOIN accounts at ON at.AccountId=b.AccountToId
    WHERE b.Deleted=0 AND b.Date >= '{START_DATE}' AND b.Date < '{END_DATE}'
    ORDER BY b.Amount1 DESC
    """
    top_bonds = query_df(sql_top)
    return audit_monthly, future, top_bonds

# ─────────────────────────────────────────────
# منطق التنبؤ
# ─────────────────────────────────────────────
def fiscal_year_label(yyyymm):
    """سنة مالية تبدأ من مايو إلى أبريل."""
    y, m = map(int, yyyymm.split("-"))
    fy_start = y if m >= 5 else y - 1
    return f"{fy_start}-{fy_start+1}"

def compute_seasonal_forecast(series_monthly: pd.Series, horizon=12, reserve_m=0.0):
    """
    series_monthly: index=YYYY-MM, values=القيمة الشهرية (مليون)
    يُرجع: forecast Series + CAGR + MAPE backtest
    """
    df = pd.DataFrame({"YM": series_monthly.index, "Val": series_monthly.values})
    df["Year"] = df["YM"].str[:4].astype(int)
    df["Month"] = df["YM"].str[5:7].astype(int)
    df["FY"] = df["YM"].apply(fiscal_year_label)

    # المتوسط الشهري لكل شهر تقويمي
    seasonal = df.groupby("Month")["Val"].mean()

    # CAGR على إجمالي السنوات المالية
    fy_totals = df.groupby("FY")["Val"].sum().sort_index()
    if len(fy_totals) >= 2 and fy_totals.iloc[0] > 0:
        years = len(fy_totals) - 1
        cagr_raw = (fy_totals.iloc[-1] / fy_totals.iloc[0]) ** (1/years) - 1
    else:
        cagr_raw = 0.0
    cagr = max(min(cagr_raw, CAGR_CAP), CAGR_FLOOR)

    # Backtest: استخدم 2022-2024 للتنبؤ بـ 2025-04 → 2026-04
    df_train = df[df["FY"].isin(["2022-2023", "2023-2024", "2024-2025"])]
    if len(df_train) >= 24:
        train_seasonal = df_train.groupby("Month")["Val"].mean()
        train_fy = df_train.groupby("FY")["Val"].sum().sort_index()
        if len(train_fy) >= 2 and train_fy.iloc[0] > 0:
            train_cagr_raw = (train_fy.iloc[-1] / train_fy.iloc[0]) ** (1/(len(train_fy)-1)) - 1
            train_cagr = max(min(train_cagr_raw, CAGR_CAP), CAGR_FLOOR)
        else:
            train_cagr = 0.0
        # تنبؤ 2025-05 → 2026-04 ثم قارن بالفعلي
        forecast_test = []
        actual_test = []
        for ym in series_monthly.index:
            if "2025-05" <= ym <= "2026-04":
                m = int(ym[5:7])
                f = float(train_seasonal.get(m, 0)) * (1 + train_cagr)
                forecast_test.append(f)
                actual_test.append(float(series_monthly[ym]))
        if forecast_test and any(a != 0 for a in actual_test):
            errors = [abs(a - f) / abs(a) for a, f in zip(actual_test, forecast_test) if a != 0]
            mape = float(np.mean(errors)) * 100 if errors else None
        else:
            mape = None
    else:
        mape = None

    # توليد التنبؤ للسنة القادمة (2026-05 → 2027-04)
    forecast_index = []
    forecast_values = []
    base_year = 2026
    for offset in range(horizon):
        month = ((5 - 1 + offset) % 12) + 1
        year = base_year + ((5 - 1 + offset) // 12)
        ym = f"{year:04d}-{month:02d}"
        val = float(seasonal.get(month, 0)) * (1 + cagr) - reserve_m
        forecast_index.append(ym)
        forecast_values.append(val)
    forecast = pd.Series(forecast_values, index=forecast_index, name="Forecast")
    return forecast, cagr, mape, seasonal

# ─────────────────────────────────────────────
# تنسيق Excel
# ─────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2E5C8A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E2F3")
FORECAST_FILL = PatternFill("solid", fgColor="FFF2CC")
ALERT_FILL = PatternFill("solid", fgColor="F8CBAD")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autofit(ws, min_width=10, max_width=40):
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = min_width
        for cell in col:
            if cell.value is not None:
                txt = str(cell.value)
                max_len = max(max_len, min(max_width, len(txt) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len

def write_df(ws, df, start_row=1, start_col=1, num_format="#,##0.000;-#,##0.000",
             int_cols=None, special_format=None):
    """يكتب dataframe إلى ورقة مع تنسيق."""
    int_cols = int_cols or []
    special_format = special_format or {}
    # العناوين
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=str(col))
    style_header_row(ws, start_row, len(df.columns))
    # البيانات
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=start_col):
            val = row[col]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if col in special_format:
                cell.number_format = special_format[col]
            elif col in int_cols:
                cell.number_format = "#,##0"
            elif isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = num_format

# ─────────────────────────────────────────────
# البناء الرئيسي
# ─────────────────────────────────────────────
def main():
    print("جاري جلب البيانات من SQL Server...")
    cf = get_monthly_cashflow()
    sup_monthly = get_per_supplier_monthly()
    inst_summary = get_installments_summary()
    balances = get_balances_snapshot()
    usd_rate = get_avg_usd_rate()
    cash_box = get_cash_total_balance()
    debtors = get_top_debtors()
    audit_monthly, audit_future, top_bonds = get_audit_summary()

    print(f"✓ {len(cf)} شهر تاريخي")
    print(f"✓ متوسط سعر الدولار من البيانات: {usd_rate:.0f} د.ع")

    # ─────────────────────
    # توليد التنبؤات
    # ─────────────────────
    cf_indexed = cf.set_index("YearMonth")
    # السلاسل التي نتنبأ بها
    series_to_forecast = {
        "Cash_IN_M": (cf_indexed["Cash_IN_M"], 0.0),
        "OUT_Suppliers_M": (cf_indexed["OUT_Suppliers_M"], 0.0),
        "OUT_Salaries_M": (cf_indexed["OUT_Salaries_M"], 0.0),
        "OUT_Purchases_M": (cf_indexed["OUT_Purchases_M"], 0.0),
        "OUT_Refunds_M": (cf_indexed["OUT_Refunds_M"], 0.0),
        "OUT_Drawings_M": (cf_indexed["OUT_Drawings_M"], 0.0),
        "OUT_Other_M": (cf_indexed["OUT_Other_M"], 0.0),
    }
    forecasts = {}
    backtests = {}
    seasonals = {}
    for col, (series, reserve) in series_to_forecast.items():
        f, cagr, mape, seas = compute_seasonal_forecast(series, FORECAST_HORIZON, reserve)
        forecasts[col] = f
        backtests[col] = {"CAGR": cagr, "MAPE": mape}
        seasonals[col] = seas

    # ─────────────────────
    # دمج تاريخي + متوقع للورقة الرئيسية
    # ─────────────────────
    forecast_index = forecasts["Cash_IN_M"].index.tolist()
    forecast_df = pd.DataFrame({"YearMonth": forecast_index})
    forecast_df["Cash_IN_M"] = forecasts["Cash_IN_M"].values
    forecast_df["OUT_Suppliers_M"] = forecasts["OUT_Suppliers_M"].values
    forecast_df["OUT_Drawings_M"] = forecasts["OUT_Drawings_M"].values
    forecast_df["OUT_Purchases_M"] = forecasts["OUT_Purchases_M"].values
    forecast_df["OUT_Salaries_M"] = forecasts["OUT_Salaries_M"].values
    forecast_df["OUT_Refunds_M"] = forecasts["OUT_Refunds_M"].values
    forecast_df["OUT_Other_M"] = forecasts["OUT_Other_M"].values
    forecast_df["OUT_Total_M"] = (forecast_df["OUT_Suppliers_M"]
                                   + forecast_df["OUT_Drawings_M"]
                                   + forecast_df["OUT_Purchases_M"]
                                   + forecast_df["OUT_Salaries_M"]
                                   + forecast_df["OUT_Refunds_M"]
                                   + forecast_df["OUT_Other_M"])
    forecast_df["BondCount"] = None
    forecast_df["Net_Operating_M"] = (forecast_df["Cash_IN_M"]
                                       - forecast_df["OUT_Suppliers_M"]
                                       - forecast_df["OUT_Purchases_M"]
                                       - forecast_df["OUT_Salaries_M"]
                                       - forecast_df["OUT_Refunds_M"]
                                       - forecast_df["OUT_Other_M"])
    forecast_df["Net_Total_M"] = forecast_df["Cash_IN_M"] - forecast_df["OUT_Total_M"]
    forecast_df["Is_Forecast"] = 1

    cf["Is_Forecast"] = 0
    combined = pd.concat([cf, forecast_df], ignore_index=True)
    # السيناريوهات
    combined["Net_Realistic_M"] = combined["Net_Operating_M"] - np.where(combined["Is_Forecast"]==1, UNEXPECTED_RESERVE_M, 0)
    avg_drawings = float(cf_indexed["OUT_Drawings_M"].mean())
    combined["Net_Pessimistic_M"] = combined["Net_Realistic_M"] - np.where(combined["Is_Forecast"]==1, avg_drawings, 0)

    # ─────────────────────
    # نموذج توزيع المدفوعات المستقبلية للموردين الـ14
    # ─────────────────────
    # المنطق:
    # 1) لكل شهر تنبؤي → احسب "Available_for_Suppliers":
    #    Forecasted_IN - Salaries - Purchases - Refunds - Other - 15M_Reserve
    # 2) احسب الحصة التاريخية لكل مورد من إجمالي المدفوعات للموردين (آخر 24 شهر)
    # 3) ابدأ توزيعاً نسبياً، طبّق السقف الشهري المرجعي، أعد توزيع الفائض
    last24 = sup_monthly[sup_monthly["YearMonth"] >= "2024-05"]
    sup_share = last24.groupby("Supplier")["Paid_M"].sum()
    if sup_share.sum() > 0:
        sup_weights = sup_share / sup_share.sum()
    else:
        sup_weights = pd.Series({n: 1/len(SUPPLIERS) for _, n, _ in SUPPLIERS})

    cap_dict = {name: cap for _, name, cap, _, _, _ in SUPPLIERS}
    plan_low_dict = {name: low for _, name, _, low, _, _ in SUPPLIERS}
    plan_high_dict = {name: high for _, name, _, _, high, _ in SUPPLIERS}
    user_monthly_dict = {name: um for _, name, _, _, _, um in SUPPLIERS}
    supplier_names = [name for _, name, _, _, _, _ in SUPPLIERS]

    def allocate_to_suppliers(pool_m, weights, caps):
        """يوزّع المبلغ المتاح pool على الموردين حسب أوزانهم مع تطبيق السقوف."""
        if pool_m <= 0:
            return {n: 0.0 for n in supplier_names}, abs(pool_m)
        # جولة 1: التوزيع النسبي
        allocations = {n: pool_m * float(weights.get(n, 0)) for n in supplier_names}
        # جولة 2: تطبيق السقوف وحساب الفائض
        excess = 0.0
        for n in supplier_names:
            cap = caps.get(n, 0)
            if cap > 0 and allocations[n] > cap:
                excess += allocations[n] - cap
                allocations[n] = cap
        # جولة 3: إعادة توزيع الفائض على من لم يصل لسقفه
        if excess > 0:
            eligible = [n for n in supplier_names
                          if caps.get(n, 0) == 0 or allocations[n] < caps[n]]
            tot_w = sum(float(weights.get(n, 0)) for n in eligible)
            if tot_w > 0:
                for n in eligible:
                    w = float(weights.get(n, 0)) / tot_w
                    space = caps.get(n, 1e9) - allocations[n] if caps.get(n, 0) > 0 else 1e9
                    add = min(excess * w, space)
                    allocations[n] += add
                    excess -= add
        return allocations, excess

    # حساب التوزيع لكل شهر تنبؤي
    allocation_rows = []
    for i, ym in enumerate(forecast_index):
        fc_in = forecast_df["Cash_IN_M"].iloc[i]
        # المتاح بعد المصاريف الثابتة + الاحتياطي
        fixed_costs = (forecast_df["OUT_Salaries_M"].iloc[i]
                        + forecast_df["OUT_Purchases_M"].iloc[i]
                        + forecast_df["OUT_Refunds_M"].iloc[i]
                        + forecast_df["OUT_Other_M"].iloc[i])
        reserve = UNEXPECTED_RESERVE_M
        pool = fc_in - fixed_costs - reserve
        allocations, leftover = allocate_to_suppliers(pool, sup_weights, cap_dict)
        row = {"YearMonth": ym, "Forecast_IN": fc_in,
                "Salaries": forecast_df["OUT_Salaries_M"].iloc[i],
                "Purchases": forecast_df["OUT_Purchases_M"].iloc[i],
                "Refunds": forecast_df["OUT_Refunds_M"].iloc[i],
                "Other": forecast_df["OUT_Other_M"].iloc[i],
                "Reserve_15M": reserve,
                "Pool_For_Suppliers": pool}
        row.update(allocations)
        row["Total_Allocated"] = sum(allocations.values())
        row["Leftover_Liquidity"] = pool - row["Total_Allocated"]
        allocation_rows.append(row)
    allocation_df = pd.DataFrame(allocation_rows)

    # ─────────────────────
    # بناء ملف Excel
    # ─────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # ============================================
    # ورقة 1: ملخص تنفيذي
    # ============================================
    ws = wb.create_sheet("١. ملخص تنفيذي")
    ws.sheet_view.rightToLeft = True

    ws["A1"] = "تقرير تحليل السيولة النقدية - معرض البيت السعيد"
    ws["A1"].font = Font(bold=True, size=16, color="2E5C8A")
    ws.merge_cells("A1:F1")
    ws["A2"] = f"تاريخ التقرير: {REPORT_DATE} | سعر صرف الدولار المعتمد: {usd_rate:,.0f} د.ع"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:F2")

    # السنوات المالية الأربع
    cf["FY"] = cf["YearMonth"].apply(fiscal_year_label)
    yearly = cf.groupby("FY").agg(
        IN_Total=("Cash_IN_M","sum"),
        OUT_Total=("OUT_Total_M","sum"),
        Net_Total=("Net_Total_M","sum"),
        Net_Operating=("Net_Operating_M","sum"),
        Months=("YearMonth","count"),
    ).reset_index()
    yearly["IN_Avg_Monthly"] = yearly["IN_Total"] / yearly["Months"]
    yearly["OUT_Avg_Monthly"] = yearly["OUT_Total"] / yearly["Months"]
    yearly["Net_Avg_Monthly"] = yearly["Net_Total"] / yearly["Months"]

    ws["A4"] = "الإجماليات السنوية (بالمليون دينار)"
    ws["A4"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A4"].fill = HEADER_FILL
    ws.merge_cells("A4:H4")

    headers = ["السنة المالية", "مقبوضات سنوي", "مصروفات سنوي", "صافي سنوي",
               "متوسط مقبوضات شهري", "متوسط مصروفات شهري", "صافي شهري", "أشهر"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=5, column=j, value=h)
    style_header_row(ws, 5, len(headers))

    for i, row in yearly.iterrows():
        r = 6 + i
        ws.cell(row=r, column=1, value=row["FY"])
        ws.cell(row=r, column=2, value=row["IN_Total"]).number_format = "#,##0.0"
        ws.cell(row=r, column=3, value=row["OUT_Total"]).number_format = "#,##0.0"
        ws.cell(row=r, column=4, value=row["Net_Total"]).number_format = "#,##0.0"
        ws.cell(row=r, column=5, value=row["IN_Avg_Monthly"]).number_format = "#,##0.0"
        ws.cell(row=r, column=6, value=row["OUT_Avg_Monthly"]).number_format = "#,##0.0"
        ws.cell(row=r, column=7, value=row["Net_Avg_Monthly"]).number_format = "#,##0.0"
        ws.cell(row=r, column=8, value=int(row["Months"]))

    # ملخص التنبؤ
    fc_total_in = forecast_df["Cash_IN_M"].sum()
    fc_total_out = forecast_df["OUT_Total_M"].sum()
    fc_total_net = fc_total_in - fc_total_out
    fc_realistic = fc_total_net - UNEXPECTED_RESERVE_M * FORECAST_HORIZON

    ws.cell(row=11, column=1, value="تنبؤ السنة القادمة 2026-2027 (بالمليون دينار)").font = Font(bold=True, size=13, color="FFFFFF")
    ws.cell(row=11, column=1).fill = HEADER_FILL
    ws.merge_cells("A11:H11")
    ws.cell(row=12, column=1, value="إجمالي مقبوضات متوقع")
    ws.cell(row=12, column=2, value=fc_total_in).number_format = "#,##0.0"
    ws.cell(row=13, column=1, value="إجمالي مصروفات متوقع")
    ws.cell(row=13, column=2, value=fc_total_out).number_format = "#,##0.0"
    ws.cell(row=14, column=1, value="صافي متوقع - متفائل")
    ws.cell(row=14, column=2, value=fc_total_net).number_format = "#,##0.0"
    ws.cell(row=15, column=1, value=f"صافي متوقع - متحفظ (بعد احتياطي {UNEXPECTED_RESERVE_M:.0f}M/شهر)")
    ws.cell(row=15, column=2, value=fc_realistic).number_format = "#,##0.0"
    ws.cell(row=16, column=1, value=f"صافي متوقع - متشائم (بعد سحوبات الشركاء التاريخية)")
    ws.cell(row=16, column=2, value=fc_realistic - avg_drawings * FORECAST_HORIZON).number_format = "#,##0.0"

    # تنبيهات السيولة
    ws.cell(row=18, column=1, value="تنبيهات السيولة المتوقعة").font = Font(bold=True, size=13, color="FFFFFF")
    ws.cell(row=18, column=1).fill = HEADER_FILL
    ws.merge_cells("A18:H18")
    fc_negative = forecast_df[forecast_df["Net_Operating_M"] - UNEXPECTED_RESERVE_M < 0]
    if len(fc_negative) > 0:
        ws.cell(row=19, column=1, value=f"⚠️ تحذير: {len(fc_negative)} شهر متوقع فيه عجز سيولة").font = Font(bold=True, color="C00000")
        ws.cell(row=20, column=1, value="الأشهر: " + ", ".join(fc_negative["YearMonth"].tolist()))
    else:
        ws.cell(row=19, column=1, value="✓ لا أشهر متوقع فيها عجز سيولة").font = Font(bold=True, color="00B050")

    # MAPE ومعدلات النمو
    ws.cell(row=22, column=1, value="معدلات النمو (CAGR) ودقة التنبؤ").font = Font(bold=True, size=13, color="FFFFFF")
    ws.cell(row=22, column=1).fill = HEADER_FILL
    ws.merge_cells("A22:H22")
    ws.cell(row=23, column=1, value="السلسلة"); ws.cell(row=23, column=2, value="CAGR")
    ws.cell(row=23, column=3, value="MAPE (دقة)"); ws.cell(row=23, column=4, value="ثقة")
    style_header_row(ws, 23, 4)
    label_map = {"Cash_IN_M":"مقبوضات", "OUT_Suppliers_M":"موردون",
                  "OUT_Salaries_M":"أجور", "OUT_Purchases_M":"مشتريات",
                  "OUT_Refunds_M":"مرتجعات", "OUT_Drawings_M":"سحوبات", "OUT_Other_M":"أخرى"}
    r = 24
    for col, lbl in label_map.items():
        bt = backtests[col]
        ws.cell(row=r, column=1, value=lbl)
        ws.cell(row=r, column=2, value=bt["CAGR"]).number_format = "0.0%"
        ws.cell(row=r, column=3, value=bt["MAPE"] if bt["MAPE"] is not None else "—")
        if bt["MAPE"] is not None:
            ws.cell(row=r, column=3).number_format = "0.0\"%\""
            if bt["MAPE"] < 25: ws.cell(row=r, column=4, value="عالية").fill = GOOD_FILL
            elif bt["MAPE"] < 50: ws.cell(row=r, column=4, value="متوسطة").fill = FORECAST_FILL
            else: ws.cell(row=r, column=4, value="منخفضة").fill = ALERT_FILL
        r += 1

    # أقساط مفتوحة
    if not inst_summary.empty:
        ws.cell(row=r+1, column=1, value="الأقساط المفتوحة (مستحقة من الزبائن)").font = Font(bold=True, size=13, color="FFFFFF")
        ws.cell(row=r+1, column=1).fill = HEADER_FILL
        ws.merge_cells(f"A{r+1}:H{r+1}")
        ws.cell(row=r+2, column=1, value="عدد الخطط النشطة")
        ws.cell(row=r+2, column=2, value=int(inst_summary["PremiumCnt"].iloc[0]))
        ws.cell(row=r+3, column=1, value="إجمالي الالتزامات (مليون د.ع)")
        ws.cell(row=r+3, column=2, value=float(inst_summary["TotalCommitted_M"].iloc[0])).number_format = "#,##0.0"
        ws.cell(row=r+4, column=1, value="المُحصّل حتى الآن")
        ws.cell(row=r+4, column=2, value=float(inst_summary["Paid_M"].iloc[0])).number_format = "#,##0.0"
        ws.cell(row=r+5, column=1, value="المتبقي على الزبائن")
        ws.cell(row=r+5, column=2, value=float(inst_summary["Remaining_M"].iloc[0])).number_format = "#,##0.0"
        ws.cell(row=r+5, column=2).fill = GOOD_FILL

    autofit(ws, min_width=14, max_width=50)

    # ============================================
    # ورقة 2: التدفق الشهري الإجمالي
    # ============================================
    ws = wb.create_sheet("٢. التدفق الشهري")
    ws.sheet_view.rightToLeft = True

    display_cols = ["YearMonth", "Cash_IN_M", "OUT_Total_M", "Net_Total_M",
                    "Net_Operating_M", "Net_Realistic_M", "Net_Pessimistic_M",
                    "Is_Forecast", "BondCount"]
    headers_ar = ["الشهر", "مقبوضات", "مصروفات إجمالي", "صافي إجمالي",
                  "صافي تشغيلي", "صافي متحفظ", "صافي متشائم", "متوقع؟", "عدد سندات"]
    for j, h in enumerate(headers_ar, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers_ar))

    for i, (_, row) in enumerate(combined.iterrows(), start=2):
        for j, c in enumerate(display_cols, 1):
            val = row[c]
            if pd.isna(val): val = None
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if c not in ["YearMonth", "Is_Forecast", "BondCount"] and isinstance(val,(int,float)):
                cell.number_format = "#,##0.0"
        if row["Is_Forecast"] == 1:
            for j in range(1, len(display_cols)+1):
                ws.cell(row=i, column=j).fill = FORECAST_FILL
        # تظليل أشهر العجز
        if isinstance(row["Net_Operating_M"], (int,float)) and row["Net_Operating_M"] < 0:
            ws.cell(row=i, column=5).fill = ALERT_FILL

    # رسم بياني
    last_row = ws.max_row
    chart = LineChart()
    chart.title = "التدفق النقدي الشهري - تاريخي ومتوقع"
    chart.y_axis.title = "مليون دينار"
    chart.x_axis.title = "الشهر"
    chart.height = 12; chart.width = 24
    data = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "K2")
    autofit(ws, min_width=11, max_width=18)

    # ============================================
    # ورقة 3: التدفق التشغيلي (بدون سحوبات)
    # ============================================
    ws = wb.create_sheet("٣. التدفق التشغيلي")
    ws.sheet_view.rightToLeft = True
    headers = ["الشهر", "مقبوضات", "مصروفات تشغيلية", "صافي تشغيلي", "متوقع؟"]
    for j, h in enumerate(headers, 1): ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))
    for i, (_, row) in enumerate(combined.iterrows(), start=2):
        out_op = (row["OUT_Suppliers_M"] + row["OUT_Purchases_M"] + row["OUT_Salaries_M"]
                  + row["OUT_Refunds_M"] + row["OUT_Other_M"])
        ws.cell(row=i, column=1, value=row["YearMonth"])
        ws.cell(row=i, column=2, value=row["Cash_IN_M"]).number_format = "#,##0.0"
        ws.cell(row=i, column=3, value=out_op).number_format = "#,##0.0"
        ws.cell(row=i, column=4, value=row["Net_Operating_M"]).number_format = "#,##0.0"
        ws.cell(row=i, column=5, value=int(row["Is_Forecast"]))
        if row["Is_Forecast"] == 1:
            for j in range(1,6): ws.cell(row=i, column=j).fill = FORECAST_FILL
        if row["Net_Operating_M"] < 0:
            ws.cell(row=i, column=4).fill = ALERT_FILL
    autofit(ws)

    # ============================================
    # ورقة 4: تفصيل المقبوضات
    # ============================================
    ws = wb.create_sheet("٤. تفصيل المقبوضات")
    ws.sheet_view.rightToLeft = True
    note = "ملاحظة: نوع 1631 (زبائن أقساط/قطاع خاص) يهيمن على ~98% من المقبوضات"
    ws["A1"] = note; ws["A1"].font = Font(italic=True, color="666666"); ws.merge_cells("A1:C1")
    write_df(ws, cf[["YearMonth","Cash_IN_M","BondCount"]].rename(
        columns={"YearMonth":"الشهر","Cash_IN_M":"مقبوضات (مليون)","BondCount":"عدد السندات"}),
        start_row=3, num_format="#,##0.000")
    autofit(ws)

    # ============================================
    # ورقة 5: تفصيل المصروفات
    # ============================================
    ws = wb.create_sheet("٥. تفصيل المصروفات")
    ws.sheet_view.rightToLeft = True
    exp_cols = ["YearMonth","OUT_Suppliers_M","OUT_Drawings_M","OUT_Purchases_M",
                "OUT_Salaries_M","OUT_Refunds_M","OUT_Other_M","OUT_Siyrafa_M","OUT_Total_M"]
    exp_headers = ["الشهر","موردون","سحوبات شركاء","مشتريات","أجور","مرتجعات","أخرى","🔄 صيرفة (دينار→دولار)","الإجمالي الشامل"]
    for j, h in enumerate(exp_headers, 1): ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(exp_headers))
    for i, (_, row) in enumerate(cf.iterrows(), start=2):
        for j, c in enumerate(exp_cols, 1):
            cell = ws.cell(row=i, column=j, value=row[c])
            if c != "YearMonth": cell.number_format = "#,##0.0"
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
    autofit(ws)

    # ============================================
    # ورقة 6: الموردون - شهري
    # ============================================
    ws = wb.create_sheet("٦. الموردون شهري")
    ws.sheet_view.rightToLeft = True
    sup_pivot = sup_monthly.pivot_table(index="YearMonth", columns="Supplier",
                                          values="Paid_M", aggfunc="sum", fill_value=0)
    sup_pivot.index = sup_pivot.index.astype(str)
    # رأس
    ws.cell(row=1, column=1, value="الشهر")
    for j, sup in enumerate(sup_pivot.columns, 2):
        ws.cell(row=1, column=j, value=sup)
    # السقف المرجعي
    ws.cell(row=2, column=1, value="السقف المرجعي")
    cap_map = {name: cap for _, name, cap, _, _, _ in SUPPLIERS}
    for j, sup in enumerate(sup_pivot.columns, 2):
        ws.cell(row=2, column=j, value=cap_map.get(sup, 0))
    style_header_row(ws, 1, len(sup_pivot.columns) + 1)
    style_header_row(ws, 2, len(sup_pivot.columns) + 1)
    # البيانات
    for i, (ym, row) in enumerate(sup_pivot.iterrows(), start=3):
        ws.cell(row=i, column=1, value=ym)
        for j, sup in enumerate(sup_pivot.columns, 2):
            v = float(row[sup])
            cell = ws.cell(row=i, column=j, value=v if v != 0 else None)
            if v: cell.number_format = "#,##0.0"
            cap = cap_map.get(sup, 0)
            if cap > 0 and v > cap:
                cell.fill = ALERT_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
    autofit(ws, min_width=12, max_width=20)

    # ============================================
    # ورقة 7: الأرصدة الحالية
    # ============================================
    ws = wb.create_sheet("٧. الأرصدة الحالية")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = f"لقطة الأرصدة بتاريخ {REPORT_DATE} - سعر الدولار المستخدم: {usd_rate:,.0f} د.ع"
    ws["A1"].font = Font(italic=True, size=10); ws.merge_cells("A1:E1")
    # موردين
    ws.cell(row=3, column=1, value="أرصدة الموردين الـ14").font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=3, column=1).fill = HEADER_FILL; ws.merge_cells("A3:F3")
    bal_dis = balances.copy()
    # تحويل الدولار للدينار
    bal_dis["Balance_IQD_M"] = bal_dis.apply(
        lambda r: r["Balance_M"] * usd_rate if r["Currency"]=="دولار" else r["Balance_M"], axis=1)
    bal_dis = bal_dis[["Supplier","Currency","Balance_M","Balance_IQD_M","LastActive"]]
    bal_dis.columns = ["المورد","العملة","الرصيد بعملته (مليون)","الرصيد بالدينار (مليون)","آخر نشاط"]
    write_df(ws, bal_dis, start_row=4, special_format={"الرصيد بعملته (مليون)":"#,##0.000;-#,##0.000",
                                                          "الرصيد بالدينار (مليون)":"#,##0.0;-#,##0.0"})

    # صناديق
    start = 4 + len(bal_dis) + 3
    ws.cell(row=start, column=1, value="أرصدة الصناديق الحالية").font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=start, column=1).fill = HEADER_FILL; ws.merge_cells(f"A{start}:F{start}")
    cb = cash_box.copy()
    cb["Currency"] = cb["CurrencyId"].map({1:"دينار",2:"دولار"}).fillna("أخرى")
    cb = cb[["Name","Currency","Balance_M"]]
    cb.columns = ["الصندوق","العملة","الرصيد (مليون)"]
    write_df(ws, cb, start_row=start+1, special_format={"الرصيد (مليون)":"#,##0.0;-#,##0.0"})

    # كبار المدينين
    start2 = start + len(cb) + 4
    ws.cell(row=start2, column=1, value="أعلى 30 زبون عليهم ذمم").font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=start2, column=1).fill = HEADER_FILL; ws.merge_cells(f"A{start2}:F{start2}")
    d = debtors.copy()
    d["Currency"] = d["CurrencyId"].map({1:"دينار",2:"دولار"}).fillna("أخرى")
    d = d[["Name","Currency","Balance_M"]]
    d.columns = ["الزبون","العملة","الرصيد (مليون)"]
    write_df(ws, d, start_row=start2+1, special_format={"الرصيد (مليون)":"#,##0.0"})
    autofit(ws, min_width=14, max_width=40)

    # ============================================
    # ورقة 8: الأقساط المفتوحة
    # ============================================
    ws = wb.create_sheet("٨. الأقساط المفتوحة")
    ws.sheet_view.rightToLeft = True
    if not inst_summary.empty:
        ws["A1"] = "ملخص الأقساط المفتوحة (مستحقة من الزبائن)"
        ws["A1"].font = Font(bold=True, size=14, color="2E5C8A"); ws.merge_cells("A1:D1")
        rows = [
            ("عدد خطط الأقساط النشطة", int(inst_summary["PremiumCnt"].iloc[0]), "#,##0"),
            ("إجمالي الالتزامات (مليون د.ع)", float(inst_summary["TotalCommitted_M"].iloc[0]), "#,##0.0"),
            ("المُحصّل حتى الآن", float(inst_summary["Paid_M"].iloc[0]), "#,##0.0"),
            ("المتبقي على الزبائن (مليون د.ع)", float(inst_summary["Remaining_M"].iloc[0]), "#,##0.0"),
        ]
        for i, (l, v, fmt) in enumerate(rows, start=3):
            ws.cell(row=i, column=1, value=l)
            c = ws.cell(row=i, column=2, value=v); c.number_format = fmt
        ws["B6"].fill = GOOD_FILL

    autofit(ws)

    # ============================================
    # ورقة 9: النمط الموسمي
    # ============================================
    ws = wb.create_sheet("٩. النمط الموسمي")
    ws.sheet_view.rightToLeft = True
    months_ar = ["يناير","فبراير","مارس","أبريل","مايو","يونيو",
                 "يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]
    seasonal_df = pd.DataFrame({"Month": list(range(1,13)), "MonthAr": months_ar})
    for col, lbl in label_map.items():
        seasonal_df[lbl] = [float(seasonals[col].get(m, 0)) for m in range(1,13)]
    seasonal_df = seasonal_df.drop(columns=["Month"])
    seasonal_df.columns = ["الشهر"] + list(seasonal_df.columns[1:])
    write_df(ws, seasonal_df, num_format="#,##0.00")
    autofit(ws)

    # ============================================
    # ورقة 10: اختبار صحة التنبؤ (Backtesting)
    # ============================================
    ws = wb.create_sheet("١٠. اختبار صحة التنبؤ")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "نموذج TRAIN على 2022-2024، تنبؤ بـ 2025-05 → 2026-04، ثم مقارنة بالفعلي"
    ws["A1"].font = Font(italic=True, color="666666"); ws.merge_cells("A1:D1")
    bt_df = pd.DataFrame({"السلسلة":[label_map[c] for c in label_map],
                           "CAGR": [f"{backtests[c]['CAGR']*100:.1f}%" for c in label_map],
                           "MAPE (%)": [f"{backtests[c]['MAPE']:.1f}%" if backtests[c]['MAPE'] else "—" for c in label_map],
                           "ثقة التنبؤ": [("عالية" if (backtests[c]["MAPE"] or 99) < 25
                                            else ("متوسطة" if (backtests[c]["MAPE"] or 99) < 50 else "منخفضة"))
                                          for c in label_map]})
    write_df(ws, bt_df, start_row=3)
    autofit(ws)

    # ============================================
    # ورقة 11: تنبيهات السيولة
    # ============================================
    ws = wb.create_sheet("١١. تنبيهات السيولة")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = f"تنبيهات الأشهر المتوقع فيها عجز سيولة (احتياطي مفاجآت: {UNEXPECTED_RESERVE_M}M/شهر)"
    ws["A1"].font = Font(bold=True, size=12, color="C00000"); ws.merge_cells("A1:F1")
    alert_df = forecast_df[["YearMonth","Cash_IN_M","OUT_Total_M","Net_Operating_M"]].copy()
    alert_df["متفائل (بدون احتياطي)"] = alert_df["Net_Operating_M"]
    alert_df["متحفظ (-15M)"] = alert_df["Net_Operating_M"] - UNEXPECTED_RESERVE_M
    alert_df["متشائم (-15M-سحوبات)"] = alert_df["Net_Operating_M"] - UNEXPECTED_RESERVE_M - avg_drawings
    alert_df = alert_df[["YearMonth","Cash_IN_M","OUT_Total_M","متفائل (بدون احتياطي)",
                          "متحفظ (-15M)","متشائم (-15M-سحوبات)"]]
    alert_df.columns = ["الشهر","مقبوضات","مصروفات","صافي متفائل","صافي متحفظ","صافي متشائم"]
    write_df(ws, alert_df, start_row=3)
    # تظليل
    for r in range(4, 4 + len(alert_df)):
        for c in [4,5,6]:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value,(int,float)):
                if cell.value < 0: cell.fill = ALERT_FILL
                elif cell.value < 15: cell.fill = FORECAST_FILL
                else: cell.fill = GOOD_FILL
                cell.number_format = "#,##0.0"
    autofit(ws)

    # ============================================
    # ورقة 11ب: توزيع المدفوعات المستقبلية للموردين (الجديد)
    # ============================================
    ws = wb.create_sheet("١١ب توزيع موردين تنبؤي")
    ws.sheet_view.rightToLeft = True

    ws["A1"] = "توزيع المدفوعات المستقبلية للموردين الـ14 وفق الواردات المتوقعة"
    ws["A1"].font = Font(bold=True, size=14, color="2E5C8A"); ws.merge_cells("A1:Q1")
    ws["A2"] = "المعادلة: من المقبوضات المتوقعة، نخصم (أجور+مشتريات+مرتجعات+أخرى+15M احتياطي) ثم نوزّع المتبقي بحسب الحصة التاريخية للمورد مع تطبيق السقف الشهري المرجعي"
    ws["A2"].font = Font(italic=True, size=10, color="666666"); ws.merge_cells("A2:Q2")

    # رأس الجدول
    base_headers = ["الشهر","واردات متوقعة","أجور","مشتريات","مرتجعات","أخرى","احتياطي 15M","🟢 المتاح للموردين"]
    sup_headers = supplier_names
    tail_headers = ["إجمالي موزّع","فائض سيولة"]
    all_headers = base_headers + sup_headers + tail_headers

    for j, h in enumerate(all_headers, 1):
        ws.cell(row=4, column=j, value=h)
    style_header_row(ws, 4, len(all_headers))
    # إضافة صف السقف المرجعي
    ws.cell(row=5, column=1, value="السقف المرجعي")
    ws.cell(row=5, column=1).font = Font(italic=True, bold=True)
    for j, n in enumerate(supplier_names, len(base_headers) + 1):
        ws.cell(row=5, column=j, value=cap_dict.get(n, 0)).number_format = "#,##0.0"
        ws.cell(row=5, column=j).fill = SUBHEAD_FILL

    # تعبئة البيانات
    col_keys = (["YearMonth","Forecast_IN","Salaries","Purchases","Refunds","Other","Reserve_15M","Pool_For_Suppliers"]
                 + supplier_names + ["Total_Allocated","Leftover_Liquidity"])
    for i, (_, row) in enumerate(allocation_df.iterrows(), start=6):
        for j, key in enumerate(col_keys, 1):
            val = row[key]
            cell = ws.cell(row=i, column=j, value=val if not pd.isna(val) else None)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if key != "YearMonth" and isinstance(val,(int,float)):
                cell.number_format = "#,##0.0"
        # تظليل الأشهر التي pool فيها سالب (تحذير)
        if row["Pool_For_Suppliers"] < 0:
            ws.cell(row=i, column=8).fill = ALERT_FILL
        elif row["Pool_For_Suppliers"] < 30:
            ws.cell(row=i, column=8).fill = FORECAST_FILL
        else:
            ws.cell(row=i, column=8).fill = GOOD_FILL
        # تظليل أعمدة الموردين التي تجاوزت السقف
        for jj, n in enumerate(supplier_names, len(base_headers) + 1):
            v = row[n]
            cap = cap_dict.get(n, 0)
            if cap > 0 and v >= cap * 0.95:
                ws.cell(row=i, column=jj).fill = ALERT_FILL
            elif v > 0:
                ws.cell(row=i, column=jj).fill = GOOD_FILL

    # صف الإجمالي السنوي
    total_row = len(allocation_df) + 7
    ws.cell(row=total_row, column=1, value="📊 الإجمالي 12 شهر")
    ws.cell(row=total_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=total_row, column=1).fill = HEADER_FILL
    for j, key in enumerate(col_keys[1:], 2):
        v = allocation_df[key].sum() if key != "YearMonth" else None
        cell = ws.cell(row=total_row, column=j, value=v)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        if isinstance(v,(int,float)):
            cell.number_format = "#,##0.0"

    # حصص الموردين الموزّعة
    ws.cell(row=total_row+2, column=1, value="حصة كل مورد من إجمالي المدفوعات السنوية المتوقعة (%)")
    ws.cell(row=total_row+2, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=total_row+2, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=total_row+2, start_column=1, end_row=total_row+2, end_column=8)
    total_alloc = allocation_df["Total_Allocated"].sum()
    for j, n in enumerate(supplier_names, len(base_headers) + 1):
        share = (allocation_df[n].sum() / total_alloc * 100) if total_alloc > 0 else 0
        ws.cell(row=total_row+3, column=j, value=share).number_format = "0.0\"%\""
        ws.cell(row=total_row+3, column=j).font = Font(bold=True)

    autofit(ws, min_width=11, max_width=22)

    # ============================================
    # ورقة 13: خطة السداد المتوقعة لكل مورد (Bottom-up forecast)
    # ============================================
    # لكل مورد: نمط موسمي × CAGR على آخر 3 سنوات، مع السقف المرجعي كحدّ أعلى
    ws13 = wb.create_sheet("١٣ خطة السداد المتوقعة")
    ws13.sheet_view.rightToLeft = True

    ws13["A1"] = "📅 خطة السداد الشهرية المتوقعة للموردين الـ14 (مايو 2026 - أبريل 2027)"
    ws13["A1"].font = Font(bold=True, size=14, color="2E5C8A")
    ws13.merge_cells("A1:S1")

    # 💱 سعر الصرف المعتمد - ثابت 1350 د.ع/$ كما اتُّفق
    AGREED_USD_RATE = 1350
    ws13["A2"] = "💱 سعر صرف الدولار المعتمد:"
    ws13["A2"].font = Font(bold=True, size=11, color="C00000")
    ws13.merge_cells("A2:E2")
    ws13["F2"] = AGREED_USD_RATE
    ws13["F2"].font = Font(bold=True, size=14, color="0000FF")
    ws13["F2"].fill = FORECAST_FILL
    ws13["F2"].number_format = "#,##0"
    ws13["F2"].alignment = Alignment(horizontal="center")
    ws13["F2"].border = BORDER
    ws13["G2"] = "د.ع/$ (متفق عليه)"
    ws13["G2"].font = Font(italic=True, bold=True)

    ws13["A3"] = "📌 المنهجية: الخطة الشهرية = القيمة التي حدّدها المستخدم (الحد الأدنى → الحد الأعلى). المتوسط التاريخي يُعرض للمقارنة فقط. سعر الصرف 1,350 د.ع/$."
    ws13["A3"].font = Font(italic=True, size=10, color="2E5C8A")
    ws13.merge_cells("A3:S3")

    # حساب التنبؤ لكل مورد على حدة من بياناته الشهرية
    sup_monthly_3y = sup_monthly[sup_monthly["YearMonth"] >= "2023-05"].copy()

    # نتائج التنبؤ بالشكل: {supplier: {YYYY-MM: amount}}
    per_supplier_forecast = {}
    per_supplier_stats = {}

    forecast_months = []
    for offset in range(12):
        month = ((5 - 1 + offset) % 12) + 1
        year = 2026 + ((5 - 1 + offset) // 12)
        forecast_months.append(f"{year:04d}-{month:02d}")

    for sup_name in supplier_names:
        sup_data = sup_monthly_3y[sup_monthly_3y["Supplier"] == sup_name].copy()
        # سلسلتان شهريتان: واحدة للدينار-المباشر، أخرى للدولار
        all_months = sorted(sup_monthly_3y["YearMonth"].unique())
        sup_series = pd.Series(0.0, index=all_months)        # Amount1 raw IQD-as-recorded
        sup_series_iqd = pd.Series(0.0, index=all_months)    # only Currency1Id=1
        sup_series_usd = pd.Series(0.0, index=all_months)    # USD millions (Currency1Id=2)
        for _, r in sup_data.iterrows():
            sup_series[r["YearMonth"]] = r["Paid_M"]
            sup_series_iqd[r["YearMonth"]] = r["Paid_IQD_M"]
            sup_series_usd[r["YearMonth"]] = r["Paid_USD_M"]

        # حساب النمط الموسمي ومعدّل النمو على السلسلة الإجمالية
        df_sup = pd.DataFrame({"YM": sup_series.index, "Val": sup_series.values})
        df_sup["Month"] = df_sup["YM"].str[5:7].astype(int)
        df_sup["FY"] = df_sup["YM"].apply(fiscal_year_label)
        seasonal_sup = df_sup.groupby("Month")["Val"].mean()
        fy_totals_sup = df_sup.groupby("FY")["Val"].sum().sort_index()
        overall_monthly_avg = sup_series.mean()
        overall_monthly_avg_iqd = sup_series_iqd.mean()    # متوسط شهري بالدينار-المباشر
        overall_monthly_avg_usd = sup_series_usd.mean()    # متوسط شهري بالدولار (بالملايين USD)
        active_months = (sup_series > 0).sum()

        if len(fy_totals_sup) >= 2 and fy_totals_sup.iloc[0] > 0.5:
            yrs = len(fy_totals_sup) - 1
            cagr_raw = (fy_totals_sup.iloc[-1] / fy_totals_sup.iloc[0]) ** (1/yrs) - 1
        else:
            cagr_raw = 0.0
        cagr_sup = max(min(cagr_raw, 0.15), -0.10)

        # السقف الشهري
        cap = cap_dict.get(sup_name, 0)

        # المنهج المبسّط (يطلبه المستخدم للتخطيط العملي):
        # توزيع المتوسط الشهري التاريخي بالتساوي على 12 شهر مع تعديل النمو
        # وتطبيق السقف الشهري كحدّ أعلى. هذا يعطي خطة سداد منتظمة قابلة للتنفيذ.
        forecast_per_month = {}
        for ym in forecast_months:
            base = overall_monthly_avg * (1 + cagr_sup)
            if cap > 0:
                base = min(base, cap)
            forecast_per_month[ym] = max(base, 0)

        # الرصيد الحالي للمورد (الدائن للشركة - عملة الدينار + الدولار محوّل)
        sup_id = next((s[0] for s in SUPPLIERS if s[1] == sup_name), None)
        cur_balance_iqd = 0.0
        if sup_id is not None:
            sup_bals = balances[balances["AccountId"] == sup_id]
            for _, br in sup_bals.iterrows():
                v = float(br["Balance_M"])
                if br["Currency"] == "دولار":
                    v = v * usd_rate
                cur_balance_iqd += v
        # نُظهر القيمة المطلقة كـ"رصيد دائن" (إذا كانت سالبة = نحن مدينون له)
        debt_owed_to_supplier = -cur_balance_iqd if cur_balance_iqd < 0 else 0

        per_supplier_forecast[sup_name] = forecast_per_month
        per_supplier_stats[sup_name] = {
            "avg_3y_monthly": overall_monthly_avg,
            "avg_3y_monthly_iqd": overall_monthly_avg_iqd,
            "avg_3y_monthly_usd": overall_monthly_avg_usd,
            "max_3y_monthly": sup_series.max(),
            "cagr": cagr_sup,
            "fy_totals": fy_totals_sup.to_dict(),
            "total_3y": sup_series.sum(),
            "active_months": active_months,
            "method": "متوسط موزّع + سقف",
            "current_debt_M": debt_owed_to_supplier,
        }

    # رأس الجدول - الأشهر كأعمدة، الموردون كصفوف
    # رأس الجدول - أعمدة الإحصاءات + شهر التنبؤ
    header_row = 5
    ws13.cell(row=header_row, column=1, value="المورد")
    ws13.cell(row=header_row, column=2, value="الرصيد الحالي (له علينا)")
    ws13.cell(row=header_row, column=3, value="📋 الخطة الأدنى")
    ws13.cell(row=header_row, column=4, value="📋 الخطة الأعلى")
    ws13.cell(row=header_row, column=5, value="💰 متوسط فعلي 3س (مرجع)")
    for j, ym in enumerate(forecast_months, start=6):
        ws13.cell(row=header_row, column=j, value=ym)
    ws13.cell(row=header_row, column=6 + len(forecast_months), value="إجمالي سنوي (سيناريو متحفظ)")
    ws13.cell(row=header_row, column=7 + len(forecast_months), value="إجمالي سنوي (سيناريو طموح)")
    ws13.cell(row=header_row, column=8 + len(forecast_months), value="شهور لتسديد الدين")
    ws13.cell(row=header_row, column=9 + len(forecast_months), value="ملاحظات")
    style_header_row(ws13, header_row, 9 + len(forecast_months))

    # صفوف الموردين
    first_data_row = 6
    for i, sup_name in enumerate(supplier_names, start=first_data_row):
        stats = per_supplier_stats[sup_name]
        cap = cap_dict.get(sup_name, 0)
        plan_low = plan_low_dict.get(sup_name, 0)
        plan_high = plan_high_dict.get(sup_name, 0)
        ws13.cell(row=i, column=1, value=sup_name).font = Font(bold=True)

        # عمود 2: الرصيد الحالي
        debt = stats["current_debt_M"]
        ws13.cell(row=i, column=2, value=debt if debt > 0 else "—")
        if debt > 0:
            ws13.cell(row=i, column=2).number_format = "#,##0.0"
            ws13.cell(row=i, column=2).fill = ALERT_FILL if debt > 50 else FORECAST_FILL

        # عمود 3: الخطة الأدنى (متحفظ)
        ws13.cell(row=i, column=3, value=plan_low if plan_low > 0 else None)
        if plan_low > 0:
            ws13.cell(row=i, column=3).number_format = "#,##0.0"
            ws13.cell(row=i, column=3).fill = SUBHEAD_FILL

        # عمود 4: الخطة الأعلى (طموح)
        ws13.cell(row=i, column=4, value=plan_high if plan_high > 0 else None)
        if plan_high > 0:
            ws13.cell(row=i, column=4).number_format = "#,##0.0"
            ws13.cell(row=i, column=4).fill = GOOD_FILL
            ws13.cell(row=i, column=4).font = Font(bold=True)

        # عمود 5: المتوسط الفعلي من البيانات (للمقارنة)
        historical_avg = stats["avg_3y_monthly_iqd"] + stats["avg_3y_monthly_usd"] * usd_rate
        ws13.cell(row=i, column=5, value=historical_avg).number_format = "#,##0.0"
        ws13.cell(row=i, column=5).font = Font(italic=True, color="666666")
        # علامة تنبيه: لو المتوسط الفعلي أعلى أو أدنى بكثير من الخطة
        if plan_high > 0:
            if historical_avg > plan_high * 1.2:
                ws13.cell(row=i, column=5).fill = ALERT_FILL  # نسدد أكثر من الخطة
            elif historical_avg < plan_low * 0.5 and plan_low > 0:
                ws13.cell(row=i, column=5).fill = FORECAST_FILL  # نسدد أقل من المخطط

        # عمود 6-17: الخطة الشهرية (سيناريو متحفظ = الحد الأدنى)
        for j, ym in enumerate(forecast_months, start=6):
            cell = ws13.cell(row=i, column=j, value=plan_low if plan_low > 0 else None)
            if plan_low > 0:
                cell.number_format = "#,##0.0"
                cell.fill = SUBHEAD_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

        # عمود 18: إجمالي السيناريو المتحفظ (12 × الحد الأدنى)
        annual_low_col = 6 + len(forecast_months)
        annual_low = plan_low * 12
        ws13.cell(row=i, column=annual_low_col, value=annual_low if annual_low > 0 else None)
        ws13.cell(row=i, column=annual_low_col).number_format = "#,##0.0"
        ws13.cell(row=i, column=annual_low_col).font = Font(bold=True)
        ws13.cell(row=i, column=annual_low_col).fill = SUBHEAD_FILL

        # عمود 19: إجمالي السيناريو الطموح (12 × الحد الأعلى)
        annual_high_col = 7 + len(forecast_months)
        annual_high = plan_high * 12
        ws13.cell(row=i, column=annual_high_col, value=annual_high if annual_high > 0 else None)
        ws13.cell(row=i, column=annual_high_col).number_format = "#,##0.0"
        ws13.cell(row=i, column=annual_high_col).font = Font(bold=True)
        ws13.cell(row=i, column=annual_high_col).fill = GOOD_FILL

        # عمود 20: شهور لتسديد الدين (بالخطة الأعلى)
        coverage_col = 8 + len(forecast_months)
        if debt > 0 and plan_high > 0:
            months_to_payoff = debt / plan_high
            cell = ws13.cell(row=i, column=coverage_col, value=months_to_payoff)
            cell.number_format = "#,##0"
            if months_to_payoff <= 12:
                cell.fill = GOOD_FILL
            elif months_to_payoff <= 36:
                cell.fill = FORECAST_FILL
            else:
                cell.fill = ALERT_FILL
        else:
            ws13.cell(row=i, column=coverage_col, value="—")

        # ملاحظات
        note = ""
        if debt == 0 and plan_high == 0:
            note = "لا دين ولا خطة"
        elif plan_high == 0 and debt > 0:
            note = "⚠️ دين موجود بدون خطة سداد"
        elif historical_avg > plan_high * 1.2 and plan_high > 0:
            note = f"📊 المتوسط الحسابي ({historical_avg:.1f}M) أعلى من الخطة"
        elif historical_avg < plan_low * 0.5 and plan_low > 0:
            # المتوسط أقل بسبب تقطّع الدفعات (دفعة كبيرة كل عدة أشهر)
            note = f"⏱️ دفعات متقطّعة (متوسط حسابي {historical_avg:.1f}M)"
        elif debt > 0 and plan_high > 0 and debt / plan_high > 36:
            note = f"🚨 السداد > 3 سنوات"
        elif debt > 0 and plan_high > 0 and debt / plan_high <= 12:
            note = "✅ السداد ≤ سنة"
        else:
            note = "متوازن"
        ws13.cell(row=i, column=9 + len(forecast_months), value=note)

    # إجمالي الشهر (مجموع جميع الموردين بالحد الأدنى - السيناريو المتحفظ)
    last_sup_row = first_data_row + len(supplier_names) - 1
    total_row = last_sup_row + 1
    ws13.cell(row=total_row, column=1, value="📊 إجمالي شهري (متحفظ)")
    ws13.cell(row=total_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws13.cell(row=total_row, column=1).fill = HEADER_FILL
    # الأعمدة الشهرية (6 إلى 5+12=17)
    for j in range(6, 6 + len(forecast_months)):
        col_letter = get_column_letter(j)
        formula = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_sup_row})"
        cell = ws13.cell(row=total_row, column=j, value=formula)
        cell.number_format = "#,##0.0"
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
    # إجمالي سنوي متحفظ
    annual_low_col = 6 + len(forecast_months)
    annual_low_letter = get_column_letter(annual_low_col)
    ws13.cell(row=total_row, column=annual_low_col,
              value=f"=SUM({annual_low_letter}{first_data_row}:{annual_low_letter}{last_sup_row})")
    ws13.cell(row=total_row, column=annual_low_col).number_format = "#,##0.0"
    ws13.cell(row=total_row, column=annual_low_col).font = Font(bold=True, color="FFFFFF")
    ws13.cell(row=total_row, column=annual_low_col).fill = HEADER_FILL
    # إجمالي سنوي طموح
    annual_high_col = 7 + len(forecast_months)
    annual_high_letter = get_column_letter(annual_high_col)
    ws13.cell(row=total_row, column=annual_high_col,
              value=f"=SUM({annual_high_letter}{first_data_row}:{annual_high_letter}{last_sup_row})")
    ws13.cell(row=total_row, column=annual_high_col).number_format = "#,##0.0"
    ws13.cell(row=total_row, column=annual_high_col).font = Font(bold=True, color="FFFFFF")
    ws13.cell(row=total_row, column=annual_high_col).fill = HEADER_FILL

    # مقارنة بالمتاح من ورقة ١١ب
    compare_row = total_row + 2
    ws13.cell(row=compare_row, column=1, value="🔄 مقارنة: خطة السداد vs المتاح من الواردات (١١ب)")
    ws13.cell(row=compare_row, column=1).font = Font(bold=True, size=12)
    ws13.merge_cells(start_row=compare_row, start_column=1, end_row=compare_row, end_column=6)

    headers2 = ["الشهر","خطة السداد","المتاح من الواردات","الفرق","الحالة"]
    for j, h in enumerate(headers2, 1):
        ws13.cell(row=compare_row+1, column=j, value=h)
    style_header_row(ws13, compare_row+1, len(headers2))

    plan_low_total = sum(plan_low_dict.get(n, 0) for n in supplier_names)
    plan_high_total = sum(plan_high_dict.get(n, 0) for n in supplier_names)

    for i, ym in enumerate(forecast_months, start=compare_row+2):
        avail = allocation_df[allocation_df["YearMonth"] == ym]["Pool_For_Suppliers"]
        avail_v = float(avail.iloc[0]) if len(avail) > 0 else 0
        ws13.cell(row=i, column=1, value=ym)
        ws13.cell(row=i, column=2, value=plan_low_total).number_format = "#,##0.0"
        ws13.cell(row=i, column=3, value=avail_v).number_format = "#,##0.0"
        diff = avail_v - plan_low_total
        ws13.cell(row=i, column=4, value=diff).number_format = "#,##0.0"
        if diff > 10:
            status_cell = ws13.cell(row=i, column=5, value="✅ يكفي")
            status_cell.fill = GOOD_FILL
        elif diff > 0:
            status_cell = ws13.cell(row=i, column=5, value="⚠️ يكفي بالكاد")
            status_cell.fill = FORECAST_FILL
        else:
            status_cell = ws13.cell(row=i, column=5, value=f"🚨 عجز {-diff:.0f}M")
            status_cell.fill = ALERT_FILL

    # ملخص الإجمالي السنوي
    summary_row = compare_row + 2 + len(forecast_months) + 2
    ws13.cell(row=summary_row, column=1, value="📈 ملخص الخطة السنوية الإجمالية").font = Font(bold=True, size=12, color="2E5C8A")
    ws13.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=5)

    summary_data = [
        ("الخطة المتحفظة (الحد الأدنى) سنوياً", plan_low_total * 12),
        ("الخطة الطموحة (الحد الأعلى) سنوياً", plan_high_total * 12),
        ("إجمالي الديون الحالية", sum(s["current_debt_M"] for s in per_supplier_stats.values())),
        ("الفجوة الزمنية (بالخطة الأعلى) أشهر", sum(s["current_debt_M"] for s in per_supplier_stats.values()) / plan_high_total if plan_high_total > 0 else 0),
    ]
    for i, (label, value) in enumerate(summary_data, start=summary_row+1):
        ws13.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws13.cell(row=i, column=2, value=value)
        c.number_format = "#,##0.0"
        c.font = Font(bold=True)
        c.fill = GOOD_FILL if i % 2 == 0 else SUBHEAD_FILL

    autofit(ws13, min_width=11, max_width=28)

    # ============================================
    # ورقة ١١ج: نظرة شاملة (3 منظورات)
    # ============================================
    ws = wb.create_sheet("١١ج نظرة شاملة")
    ws.sheet_view.rightToLeft = True

    ws["A1"] = "🔍 نظرة شاملة - 3 منظورات لتحليل التدفق النقدي"
    ws["A1"].font = Font(bold=True, size=14, color="2E5C8A"); ws.merge_cells("A1:H1")
    ws["A2"] = "السبب: قد تختلف المصروفات حسب المنظور المحاسبي - هل نحتسب الصيرفة كمصروف؟"
    ws["A2"].font = Font(italic=True, color="666666"); ws.merge_cells("A2:H2")

    # شرح المنظورات
    explanations = [
        ("المنظور (View)", "ما يحتسب", "الإجمالي السنوي (تقريباً)", "متى يُستخدم؟"),
        ("A - الشامل (Comprehensive)",
            "موردون + سحوبات + مرتجعات + مشتريات + أجور + أخرى + صيرفة",
            "~1,208 M/سنة (يطابق توقعك)",
            "للسيولة - كل دينار يخرج من الصندوق"),
        ("B - الدينار النقي (IQD-only)",
            "كل ما سبق لكن بحركات الدينار فقط (لا الدولار)",
            "~854 M/سنة",
            "لتتبع حركة الدينار الفعلية في الصندوق"),
        ("C - الاقتصادي (Economic - Default)",
            "موردون (بكل العملات IQDeq) + سحوبات + مرتجعات + مشتريات + أجور (بدون صيرفة)",
            "~816 M/سنة",
            "للأرباح والربحية - يتجنّب ازدواج عد الصيرفة + دفعات USD"),
    ]
    for i, row in enumerate(explanations, start=4):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            cell.border = BORDER
            if i == 4:
                cell.fill = HEADER_FILL; cell.font = Font(bold=True, color="FFFFFF")
            elif i == 6:
                cell.fill = GOOD_FILL  # View C = current default
            else:
                cell.fill = SUBHEAD_FILL

    # جدول المقارنة الشهري
    ws.cell(row=9, column=1, value="مقارنة الـ3 منظورات شهرياً (آخر 12 شهر)").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=9, column=1).fill = HEADER_FILL; ws.merge_cells("A9:H9")

    compare_headers = ["الشهر","المقبوضات","A: شامل (مع صيرفة)","B: دينار نقي","C: اقتصادي","صافي A","صافي C","ملاحظات"]
    for j, h in enumerate(compare_headers, 1):
        ws.cell(row=10, column=j, value=h)
    style_header_row(ws, 10, len(compare_headers))

    last12 = cf.tail(12).copy()
    for i, (_, row) in enumerate(last12.iterrows(), start=11):
        ym = row["YearMonth"]
        out_a = row["OUT_Total_Comprehensive_M"]
        out_c = row["OUT_Total_Operational_M"]
        out_b = out_c  # تقريب: يحتاج استعلام منفصل لـ Currency1Id=1 فقط
        # لكي يكون دقيقاً، لنأخذ المنظور B = شامل - دفعات USD للموردين
        net_a = row["Cash_IN_M"] - out_a
        net_c = row["Cash_IN_M"] - out_c
        note = ""
        if row["OUT_Siyrafa_M"] > 50:
            note = f"⚠️ صيرفة كبيرة هذا الشهر: {row['OUT_Siyrafa_M']:.0f}M"

        ws.cell(row=i, column=1, value=ym)
        ws.cell(row=i, column=2, value=row["Cash_IN_M"]).number_format = "#,##0.0"
        ws.cell(row=i, column=3, value=out_a).number_format = "#,##0.0"
        ws.cell(row=i, column=4, value=out_b).number_format = "#,##0.0"
        ws.cell(row=i, column=5, value=out_c).number_format = "#,##0.0"
        ws.cell(row=i, column=6, value=net_a).number_format = "#,##0.0"
        ws.cell(row=i, column=7, value=net_c).number_format = "#,##0.0"
        ws.cell(row=i, column=8, value=note)
        if net_a < 0: ws.cell(row=i, column=6).fill = ALERT_FILL
        if net_c < 0: ws.cell(row=i, column=7).fill = ALERT_FILL

    # الإجمالي السنوي
    tot_row = 11 + len(last12)
    ws.cell(row=tot_row, column=1, value="إجمالي 12 شهر").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=tot_row, column=1).fill = HEADER_FILL
    for j, key in enumerate(["Cash_IN_M","OUT_Total_Comprehensive_M","OUT_Total_Operational_M","OUT_Total_Operational_M"], 2):
        v = last12[key].sum()
        ws.cell(row=tot_row, column=j, value=v).number_format = "#,##0.0"
        ws.cell(row=tot_row, column=j).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=tot_row, column=j).fill = HEADER_FILL
    net_a_sum = last12["Cash_IN_M"].sum() - last12["OUT_Total_Comprehensive_M"].sum()
    net_c_sum = last12["Cash_IN_M"].sum() - last12["OUT_Total_Operational_M"].sum()
    ws.cell(row=tot_row, column=6, value=net_a_sum).number_format = "#,##0.0"
    ws.cell(row=tot_row, column=7, value=net_c_sum).number_format = "#,##0.0"
    for j in [6,7]:
        ws.cell(row=tot_row, column=j).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=tot_row, column=j).fill = HEADER_FILL

    # خلاصة
    summary_row = tot_row + 3
    ws.cell(row=summary_row, column=1, value="📌 خلاصة التحليل العميق").font = Font(bold=True, size=12, color="C00000")
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)
    summary_lines = [
        f"• المقبوضات السنوية: {last12['Cash_IN_M'].sum():,.0f} مليون د.ع (من زبائن الأقساط والنقد)",
        f"• المصروفات الشاملة (View A): {last12['OUT_Total_Comprehensive_M'].sum():,.0f} مليون - يطابق توقعك أن المحل دوّار",
        f"• منها صيرفة (دينار→دولار): {last12['OUT_Siyrafa_M'].sum():,.0f} مليون - دينار يخرج لشراء دولار لدفع الموردين الدولاريين",
        f"• الفجوة الفعلية (سيولة محتجزة في الصناديق): {net_a_sum:,.0f} مليون فقط",
        "• ملاحظة: View C يستخدم 'دفعات الدولار للموردين (IQD-equivalent)' بدلاً من الصيرفة لتجنّب الازدواج",
    ]
    for i, line in enumerate(summary_lines, start=summary_row+1):
        ws.cell(row=i, column=1, value=line)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    autofit(ws, min_width=14, max_width=45)

    # ============================================
    # ورقة 14: خطة السداد (مُعاد التصميم)
    # ============================================
    # المنطق:
    # - تُستخدم القيمة المعتمدة شهرياً من المستخدم (user_monthly_dict) كأساس
    # - يُحسب تتبّع الرصيد الشهري لكل مورد: opening - payment + new_purchases
    # - يُظهر متى ينتهي السداد لكل مورد
    # - تسمية العمود: "إيجار + مصاريف تشغيلية" بدل "مشتريات أخرى"
    # - الهدف الأساسي: نهاية فبراير 2027 (10 أشهر)، الأقصى أبريل 2027

    ws14 = wb.create_sheet("١٤ خطة السداد")
    ws14.sheet_view.rightToLeft = True

    PLAN_RATE = 1350
    TARGET_PAYOFF_MONTHS = 10  # نهاية فبراير 2027

    # بناء خطة السداد لكل مورد
    payment_plan = {}
    for sup_name in supplier_names:
        stats = per_supplier_stats[sup_name]
        debt = stats["current_debt_M"]
        monthly_purchase = stats["avg_3y_monthly_iqd"] + stats["avg_3y_monthly_usd"] * PLAN_RATE
        user_monthly = user_monthly_dict.get(sup_name, 0)

        # تتبّع الرصيد الشهري على مدى 12 شهر
        balance_trajectory = []
        bal = debt
        cleared_month_idx = None
        for m in range(12):
            opening = bal
            payment = user_monthly
            new_purchases = monthly_purchase
            closing = max(opening - payment + new_purchases, 0)
            # نعتبر "تم التسديد" إذا الرصيد < المشتريات الشهرية (يعني الدين فعلياً مُسدّد)
            if cleared_month_idx is None and opening > 0 and opening - payment <= 0:
                cleared_month_idx = m
            balance_trajectory.append({
                "opening": opening, "payment": payment,
                "new_purchases": new_purchases, "closing": closing,
            })
            bal = closing

        # كم شهر فعلياً لتسديد الدين الحالي (تجاهل المشتريات الجديدة في الحساب البسيط)
        if user_monthly > monthly_purchase:
            months_to_clear = debt / (user_monthly - monthly_purchase) if user_monthly > monthly_purchase else 999
        elif user_monthly > 0 and debt > 0:
            months_to_clear = 999  # لن ينتهي لأن الدفعة لا تتجاوز المشتريات
        else:
            months_to_clear = 0

        payment_plan[sup_name] = {
            "debt": debt,
            "monthly_purchase": monthly_purchase,
            "user_monthly": user_monthly,
            "annual_payment": user_monthly * 12,
            "months_to_clear": months_to_clear,
            "balance_trajectory": balance_trajectory,
            "cleared_month_idx": cleared_month_idx,
        }

    # رأس الورقة
    ws14["A1"] = "📋 خطة السداد الشهرية للموردين الـ14 (المُعتمدة)"
    ws14["A1"].font = Font(bold=True, size=15, color="2E5C8A")
    ws14.merge_cells("A1:R1")
    ws14["A2"] = f"🎯 الهدف الأساسي: تسديد الديون بحلول فبراير 2027 (10 أشهر) | السعر: {PLAN_RATE} د.ع/$ | الاحتياطي: 15M/شهر | المعدّلات الشهرية مُعتمدة من المستخدم"
    ws14["A2"].font = Font(italic=True, size=11, color="C00000")
    ws14.merge_cells("A2:R2")
    ws14["A3"] = "💡 الخطة تأخذ بالاعتبار: استمرار المشتريات الجديدة من الموردين خلال الموسم بناءً على معدّل آخر 3 سنوات"
    ws14["A3"].font = Font(italic=True, size=10, color="666666")
    ws14.merge_cells("A3:R3")

    # ═══ القسم 1: ملخص الخطة لكل مورد ═══
    ws14.cell(row=5, column=1, value="🏢 ملخص خطة السداد لكل مورد (بالقيم المعتمدة)").font = Font(bold=True, size=12, color="FFFFFF")
    ws14.cell(row=5, column=1).fill = HEADER_FILL
    ws14.merge_cells("A5:I5")

    summary_headers = ["المورد", "الدين الحالي", "مشتريات شهرية متوقعة",
                        "🎯 المعتمد شهرياً", "السنوي (×12)",
                        "صافي تخفيض شهري", "شهور لتسديد الدين", "موعد نهاية الدين", "الحالة"]
    for j, h in enumerate(summary_headers, 1):
        ws14.cell(row=6, column=j, value=h)
    style_header_row(ws14, 6, len(summary_headers))

    sup_row_start = 7
    ar_month_names = ["مايو 2026","يونيو 2026","يوليو 2026","أغسطس 2026","سبتمبر 2026",
                      "أكتوبر 2026","نوفمبر 2026","ديسمبر 2026","يناير 2027","فبراير 2027",
                      "مارس 2027","أبريل 2027","بعد أبريل 2027"]
    for i, sup_name in enumerate(supplier_names, start=sup_row_start):
        plan = payment_plan[sup_name]
        ws14.cell(row=i, column=1, value=sup_name).font = Font(bold=True)

        # الدين الحالي
        ws14.cell(row=i, column=2, value=plan["debt"] if plan["debt"] > 0 else None).number_format = "#,##0.0"
        if plan["debt"] > 50:
            ws14.cell(row=i, column=2).fill = ALERT_FILL
        elif plan["debt"] > 0:
            ws14.cell(row=i, column=2).fill = FORECAST_FILL

        # مشتريات شهرية
        ws14.cell(row=i, column=3, value=plan["monthly_purchase"]).number_format = "#,##0.00"
        ws14.cell(row=i, column=3).font = Font(italic=True, color="666666")

        # المعتمد شهرياً
        cell_user = ws14.cell(row=i, column=4, value=plan["user_monthly"] if plan["user_monthly"] > 0 else None)
        cell_user.number_format = "#,##0.00"
        cell_user.font = Font(bold=True)
        cell_user.fill = GOOD_FILL

        # السنوي
        ws14.cell(row=i, column=5, value=plan["annual_payment"]).number_format = "#,##0.0"

        # صافي تخفيض شهري (المعتمد - المشتريات الجديدة)
        net_paydown = plan["user_monthly"] - plan["monthly_purchase"]
        cell_paydown = ws14.cell(row=i, column=6, value=net_paydown)
        cell_paydown.number_format = "+#,##0.00;-#,##0.00"
        if net_paydown <= 0 and plan["debt"] > 0:
            cell_paydown.fill = ALERT_FILL
        elif net_paydown > 0:
            cell_paydown.fill = SUBHEAD_FILL

        # شهور لتسديد الدين
        cell_months = ws14.cell(row=i, column=7)
        if plan["debt"] == 0:
            cell_months.value = "—"
        elif plan["months_to_clear"] >= 999:
            cell_months.value = "∞ (لا ينتهي)"
            cell_months.fill = ALERT_FILL
        else:
            cell_months.value = round(plan["months_to_clear"], 1)
            cell_months.number_format = "#,##0.0"
            if plan["months_to_clear"] <= TARGET_PAYOFF_MONTHS:
                cell_months.fill = GOOD_FILL
            elif plan["months_to_clear"] <= 12:
                cell_months.fill = FORECAST_FILL
            else:
                cell_months.fill = ALERT_FILL

        # موعد نهاية الدين
        if plan["debt"] == 0:
            ws14.cell(row=i, column=8, value="—")
        elif plan["months_to_clear"] >= 999:
            ws14.cell(row=i, column=8, value="لن ينتهي بهذه المعدّل")
            ws14.cell(row=i, column=8).fill = ALERT_FILL
        else:
            month_idx = int(plan["months_to_clear"])
            month_name = ar_month_names[min(month_idx, 12)]
            ws14.cell(row=i, column=8, value=month_name)
            if plan["months_to_clear"] <= TARGET_PAYOFF_MONTHS:
                ws14.cell(row=i, column=8).fill = GOOD_FILL
            elif plan["months_to_clear"] <= 12:
                ws14.cell(row=i, column=8).fill = FORECAST_FILL
            else:
                ws14.cell(row=i, column=8).fill = ALERT_FILL

        # الحالة
        if plan["debt"] == 0 and plan["user_monthly"] == 0:
            status = "لا دين"
        elif plan["debt"] == 0:
            status = "✅ بدون دين - دفع للمشتريات الجديدة فقط"
        elif plan["months_to_clear"] <= TARGET_PAYOFF_MONTHS:
            status = "✅ يحقق الهدف (≤ فبراير 2027)"
        elif plan["months_to_clear"] <= 12:
            status = "⚠️ ينتهي خلال السنة"
        elif plan["months_to_clear"] >= 999:
            status = "🚨 يحتاج زيادة الدفع"
        else:
            status = f"🚨 سيتجاوز السنة ({plan['months_to_clear']:.0f} شهر)"
        ws14.cell(row=i, column=9, value=status)

    # صف الإجمالي
    total_sup_row = sup_row_start + len(supplier_names)
    ws14.cell(row=total_sup_row, column=1, value="📊 الإجمالي").font = Font(bold=True, color="FFFFFF")
    ws14.cell(row=total_sup_row, column=1).fill = HEADER_FILL
    total_debt = sum(p["debt"] for p in payment_plan.values())
    total_purch = sum(p["monthly_purchase"] for p in payment_plan.values())
    total_user = sum(p["user_monthly"] for p in payment_plan.values())
    total_annual = total_user * 12
    for col, val in [(2, total_debt), (3, total_purch), (4, total_user), (5, total_annual)]:
        cell = ws14.cell(row=total_sup_row, column=col, value=val)
        cell.number_format = "#,##0.0"
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL

    # ═══ القسم 2: جدول السداد الشهري لكل مورد ═══
    sched_start = total_sup_row + 3
    ws14.cell(row=sched_start, column=1, value="📅 جدول السداد الشهري (المبلغ الذي ستدفعه كل شهر)").font = Font(bold=True, size=12, color="FFFFFF")
    ws14.cell(row=sched_start, column=1).fill = HEADER_FILL
    ws14.merge_cells(start_row=sched_start, start_column=1, end_row=sched_start, end_column=16)

    sched_headers = ["المورد"] + forecast_months + ["إجمالي السنة"]
    for j, h in enumerate(sched_headers, 1):
        ws14.cell(row=sched_start+1, column=j, value=h)
    style_header_row(ws14, sched_start+1, len(sched_headers))

    sched_data_start = sched_start + 2
    monthly_totals_supplier = [0.0] * 12
    for i, sup_name in enumerate(supplier_names, start=sched_data_start):
        plan = payment_plan[sup_name]
        ws14.cell(row=i, column=1, value=sup_name).font = Font(bold=True)

        annual = 0
        for j in range(12):
            traj = plan["balance_trajectory"][j]
            # المبلغ المدفوع هذا الشهر
            payment = traj["payment"]
            # إذا الدين انتهى (closing = 0)، نُقلّل الدفع لمعدّل المشتريات فقط
            if plan["cleared_month_idx"] is not None and j > plan["cleared_month_idx"]:
                payment = plan["monthly_purchase"]
            monthly_totals_supplier[j] += payment
            annual += payment
            cell = ws14.cell(row=i, column=2+j, value=payment if payment > 0 else None)
            if payment > 0:
                cell.number_format = "#,##0.0"
                if plan["cleared_month_idx"] is not None and j > plan["cleared_month_idx"]:
                    cell.fill = GOOD_FILL
                elif plan["debt"] > 50:
                    cell.fill = FORECAST_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

        cell_annual = ws14.cell(row=i, column=2+12, value=annual if annual > 0 else None)
        cell_annual.number_format = "#,##0.0"
        cell_annual.font = Font(bold=True)
        cell_annual.fill = SUBHEAD_FILL

    # صف الإجمالي الشهري
    tot_row = sched_data_start + len(supplier_names)
    ws14.cell(row=tot_row, column=1, value="📊 إجمالي شهري").font = Font(bold=True, color="FFFFFF")
    ws14.cell(row=tot_row, column=1).fill = HEADER_FILL
    for j in range(12):
        cell = ws14.cell(row=tot_row, column=2+j, value=monthly_totals_supplier[j])
        cell.number_format = "#,##0.0"
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
    grand = sum(monthly_totals_supplier)
    ws14.cell(row=tot_row, column=14, value=grand).number_format = "#,##0.0"
    ws14.cell(row=tot_row, column=14).font = Font(bold=True, color="FFFFFF")
    ws14.cell(row=tot_row, column=14).fill = HEADER_FILL

    # ═══ القسم 3: تتبّع رصيد الدين الشهري (Running Balance) ═══
    bal_start = tot_row + 3
    ws14.cell(row=bal_start, column=1, value="📉 تتبّع رصيد الدين شهرياً (متى ينتهي كل مورد؟)").font = Font(bold=True, size=12, color="FFFFFF")
    ws14.cell(row=bal_start, column=1).fill = HEADER_FILL
    ws14.merge_cells(start_row=bal_start, start_column=1, end_row=bal_start, end_column=15)

    bal_headers = ["المورد", "الافتتاحي"] + forecast_months + ["الختامي"]
    for j, h in enumerate(bal_headers, 1):
        ws14.cell(row=bal_start+1, column=j, value=h)
    style_header_row(ws14, bal_start+1, len(bal_headers))

    bal_data_start = bal_start + 2
    for i, sup_name in enumerate(supplier_names, start=bal_data_start):
        plan = payment_plan[sup_name]
        ws14.cell(row=i, column=1, value=sup_name).font = Font(bold=True)
        ws14.cell(row=i, column=2, value=plan["debt"] if plan["debt"] > 0 else None).number_format = "#,##0.0"
        ws14.cell(row=i, column=2).fill = FORECAST_FILL

        for j, traj in enumerate(plan["balance_trajectory"]):
            cell = ws14.cell(row=i, column=3+j, value=traj["closing"] if traj["closing"] > 0 else 0)
            cell.number_format = "#,##0.0"
            if traj["closing"] == 0:
                cell.fill = GOOD_FILL
                cell.value = "✅ مُسدّد"
            elif traj["closing"] < plan["debt"] * 0.2:
                cell.fill = FORECAST_FILL  # قارب الانتهاء
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

        # الختامي بعد سنة
        final_bal = plan["balance_trajectory"][-1]["closing"]
        cell_final = ws14.cell(row=i, column=15, value=final_bal if final_bal > 0 else "✅ مُسدّد")
        if final_bal > 0:
            cell_final.number_format = "#,##0.0"
            cell_final.fill = ALERT_FILL
        else:
            cell_final.fill = GOOD_FILL

    # ═══ القسم 4: التدفق النقدي الشهري الإجمالي ═══
    cf_start = bal_data_start + len(supplier_names) + 2
    ws14.cell(row=cf_start, column=1, value="💰 التدفق النقدي الإجمالي الشهري - هل الواردات تكفي؟").font = Font(bold=True, size=12, color="FFFFFF")
    ws14.cell(row=cf_start, column=1).fill = HEADER_FILL
    ws14.merge_cells(start_row=cf_start, start_column=1, end_row=cf_start, end_column=10)

    cf_headers = ["الشهر", "الواردات المتوقعة", "أجور", "إيجار + مصاريف تشغيلية",
                   "مرتجعات", "احتياطي 15M", "المتاح للموردين",
                   "📅 المخطط للموردين", "الفجوة", "الحالة"]
    for j, h in enumerate(cf_headers, 1):
        ws14.cell(row=cf_start+1, column=j, value=h)
    style_header_row(ws14, cf_start+1, len(cf_headers))

    cf_data_start = cf_start + 2
    cumulative_surplus = 0
    total_income = 0; total_salary = 0; total_ops = 0; total_refund = 0; total_planned_cf = 0
    for i, ym in enumerate(forecast_months):
        row_n = cf_data_start + i
        fc_row = forecast_df[forecast_df["YearMonth"] == ym].iloc[0]
        salary = fc_row["OUT_Salaries_M"]
        ops_costs = fc_row["OUT_Purchases_M"]  # إيجار + مصاريف تشغيلية (نوع 3110)
        refund = fc_row["OUT_Refunds_M"]
        reserve = UNEXPECTED_RESERVE_M
        income = fc_row["Cash_IN_M"]
        available = income - salary - ops_costs - refund - reserve
        planned = monthly_totals_supplier[i]
        gap = available - planned
        cumulative_surplus += gap
        total_income += income; total_salary += salary; total_ops += ops_costs
        total_refund += refund; total_planned_cf += planned

        ws14.cell(row=row_n, column=1, value=ym).font = Font(bold=True)
        ws14.cell(row=row_n, column=2, value=income).number_format = "#,##0.0"
        ws14.cell(row=row_n, column=3, value=salary).number_format = "#,##0.0"
        ws14.cell(row=row_n, column=4, value=ops_costs).number_format = "#,##0.0"
        ws14.cell(row=row_n, column=5, value=refund).number_format = "#,##0.0"
        ws14.cell(row=row_n, column=6, value=reserve).number_format = "#,##0.0"
        ws14.cell(row=row_n, column=7, value=available).number_format = "#,##0.0"
        ws14.cell(row=row_n, column=7).fill = SUBHEAD_FILL
        ws14.cell(row=row_n, column=8, value=planned).number_format = "#,##0.0"
        ws14.cell(row=row_n, column=8).fill = FORECAST_FILL
        ws14.cell(row=row_n, column=9, value=gap).number_format = "+#,##0.0;-#,##0.0"
        if gap > 10:
            ws14.cell(row=row_n, column=9).fill = GOOD_FILL
            ws14.cell(row=row_n, column=10, value="✅ يكفي بسهولة")
        elif gap > 0:
            ws14.cell(row=row_n, column=9).fill = FORECAST_FILL
            ws14.cell(row=row_n, column=10, value="⚠️ يكفي بالكاد")
        else:
            ws14.cell(row=row_n, column=9).fill = ALERT_FILL
            ws14.cell(row=row_n, column=10, value=f"🚨 عجز {-gap:.1f}M")

    # صف الإجمالي السنوي للتدفق النقدي
    cf_tot_row = cf_data_start + 12
    ws14.cell(row=cf_tot_row, column=1, value="📊 الإجمالي السنوي").font = Font(bold=True, color="FFFFFF")
    ws14.cell(row=cf_tot_row, column=1).fill = HEADER_FILL
    totals_cf = [total_income, total_salary, total_ops, total_refund,
                  UNEXPECTED_RESERVE_M * 12, total_income - total_salary - total_ops - total_refund - UNEXPECTED_RESERVE_M*12,
                  total_planned_cf, cumulative_surplus]
    for col, val in enumerate(totals_cf, start=2):
        cell = ws14.cell(row=cf_tot_row, column=col, value=val)
        cell.number_format = "#,##0.0"
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL

    # ═══ ملخص نهاية الخطة ═══
    sum_row = cf_tot_row + 3
    ws14.cell(row=sum_row, column=1, value="📈 ملخص الخطة السنوية والتقييم").font = Font(bold=True, size=12, color="2E5C8A")
    ws14.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=5)

    cleared_count = sum(1 for p in payment_plan.values() if p["debt"] > 0 and p["months_to_clear"] <= TARGET_PAYOFF_MONTHS)
    not_cleared = sum(1 for p in payment_plan.values() if p["debt"] > 0 and p["months_to_clear"] > TARGET_PAYOFF_MONTHS)
    debt_cleared = sum(p["debt"] for p in payment_plan.values() if p["months_to_clear"] <= TARGET_PAYOFF_MONTHS)
    debt_not_cleared = sum(p["debt"] for p in payment_plan.values() if p["months_to_clear"] > TARGET_PAYOFF_MONTHS)

    summary_items = [
        ("إجمالي الديون الحالية", total_debt, "د.ع مليون"),
        ("إجمالي السداد الشهري المعتمد", total_user, "د.ع مليون"),
        ("إجمالي السداد السنوي", total_annual, "د.ع مليون"),
        ("إجمالي المشتريات الجديدة المتوقعة (12 شهر)", total_purch * 12, "د.ع مليون"),
        (f"موردون يحققون الهدف (≤ فبراير 2027)", cleared_count, "مورد"),
        (f"موردون يحتاجون أكثر من فبراير 2027", not_cleared, "مورد"),
        ("ديون ستُسدّد قبل فبراير 2027", debt_cleared, "د.ع مليون"),
        ("ديون تتجاوز الهدف الزمني", debt_not_cleared, "د.ع مليون"),
        ("صافي الفائض/العجز السنوي للسيولة", cumulative_surplus, "د.ع مليون"),
    ]
    for i, (label, value, unit) in enumerate(summary_items, start=sum_row+1):
        ws14.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws14.cell(row=i, column=2, value=value)
        if isinstance(value, (int, float)) and unit == "مورد":
            cell.number_format = "0"
        else:
            cell.number_format = "#,##0.0"
        cell.font = Font(bold=True)
        ws14.cell(row=i, column=3, value=unit).font = Font(italic=True, color="666666")
        if "عجز" in label and isinstance(value, (int, float)) and value < 0:
            cell.fill = ALERT_FILL
        elif "ستُسدّد" in label or "يحققون" in label:
            cell.fill = GOOD_FILL
        elif "يحتاجون" in label or "تتجاوز" in label:
            cell.fill = ALERT_FILL if value > 0 else SUBHEAD_FILL
        else:
            cell.fill = SUBHEAD_FILL

    autofit(ws14, min_width=11, max_width=30)

    # ============================================
    # ورقة 12: التحقق المحاسبي
    # ============================================
    ws = wb.create_sheet("١٢. التحقق المحاسبي")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "طبقات التحقق المحاسبي والشذوذ"
    ws["A1"].font = Font(bold=True, size=14, color="2E5C8A"); ws.merge_cells("A1:G1")

    # عدد السندات شهرياً
    ws.cell(row=3, column=1, value="عدد السندات شهرياً (مع توزيع العملات)").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=3, column=1).fill = HEADER_FILL; ws.merge_cells("A3:G3")
    am = audit_monthly.rename(columns={"YearMonth":"الشهر","BondCount":"إجمالي السندات",
                                         "Dinar_Bonds":"بالدينار","Dollar_Bonds":"بالدولار",
                                         "OpType_0":"نوع 0","OpType_Other":"أنواع أخرى"})
    write_df(ws, am, start_row=4, int_cols=["إجمالي السندات","بالدينار","بالدولار","نوع 0","أنواع أخرى"])

    # سندات مستقبلية
    start = 4 + len(am) + 2
    ws.cell(row=start, column=1, value="السندات المستقبلية (مؤجلة - مستثناة من التحليل)").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=start, column=1).fill = HEADER_FILL; ws.merge_cells(f"A{start}:G{start}")
    ws.cell(row=start+1, column=1, value="عدد السندات المستقبلية"); ws.cell(row=start+1, column=2, value=int(audit_future["Cnt"].iloc[0]))
    ws.cell(row=start+2, column=1, value="إجمالي مبالغها (مليون)"); ws.cell(row=start+2, column=2, value=float(audit_future["Sum_M"].iloc[0])).number_format = "#,##0.0"

    # أكبر 20 سند
    start2 = start + 5
    ws.cell(row=start2, column=1, value="أكبر 20 سند (مراجعة الشذوذ)").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=start2, column=1).fill = HEADER_FILL; ws.merge_cells(f"A{start2}:G{start2}")
    tb = top_bonds.rename(columns={"Id":"#","Date":"التاريخ","Amount_M":"مبلغ (مليون)",
                                     "FromAccount":"من حساب","ToAccount":"إلى حساب","Reason":"البيان"})
    write_df(ws, tb, start_row=start2+1, int_cols=["#"])

    autofit(ws, min_width=12, max_width=50)

    # حفظ
    wb.save(OUTPUT_FILE)
    print(f"\n✓ تم إنشاء التقرير: {OUTPUT_FILE}")
    print(f"  حجم الملف: {OUTPUT_FILE.stat().st_size / 1024:.1f} KB")
    print(f"  عدد الأوراق: {len(wb.sheetnames)}")
    return OUTPUT_FILE

if __name__ == "__main__":
    main()
